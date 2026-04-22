"""openpyxl-based Excel workbook builder for Seeker Accounting exports.

Produces .xlsx files with:
  - Company identity header rows
  - Seeker Accounting branded footer rows
  - Professional column/row styling (header colours, stripe rows, totals)
  - Page setup for A4 or A5 printing
  - Support for multiple sheets

Usage:
    builder = ExcelWorkbookBuilder(page_size=PageSize.A4)
    sheet = builder.add_sheet("Invoices")
    sheet.write_document_header(company, "Sales Invoices", generated_at="2026-03-31")
    sheet.write_table_header(["#", "Customer", "Date", "Amount"])
    sheet.write_table_row(["INV-001", "Acme Corp", "2026-03-31", "1,500,000"])
    sheet.write_totals_row(["", "", "TOTAL", "1,500,000"])
    sheet.write_branded_footer()
    builder.save("/path/to/output.xlsx")
"""
from __future__ import annotations

import os

from seeker_accounting.platform.printing.print_data_protocol import (
    CompanyHeaderData,
    PageOrientation,
    PageSize,
)

# ── Brand colours (openpyxl PatternFill / Font use hex without '#') ──────────

_C_BRAND_PRIMARY = "1E3A5F"
_C_BRAND_LIGHT = "F0F3F7"
_C_BRAND_TOTAL = "EEF2F7"
_C_STRIPE = "F9FAFB"
_C_WHITE = "FFFFFF"
_C_MUTED = "6B7280"
_C_BORDER = "D0D7DE"

_BRAND_NAME = "Seeker Accounting"
_BRAND_TAGLINE = "Built for Business Clarity. Designed For Success."

# openpyxl paper size constants
_PAPER_A4 = 9
_PAPER_A5 = 11


class ExcelSheetBuilder:
    """Builds data content on a single Excel worksheet."""

    def __init__(
        self,
        ws,  # openpyxl Worksheet
        *,
        page_size: PageSize = PageSize.A4,
        orientation: PageOrientation = PageOrientation.PORTRAIT,
    ) -> None:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.styles import numbers as xl_numbers

        self._ws = ws
        self._page_size = page_size
        self._orientation = orientation
        self._current_row = 1
        self._num_columns = 1
        self._col_widths: dict[int, float] = {}

        # Reusable style factories (stored as constructors to avoid pickling issues)
        self._Font = Font
        self._Fill = PatternFill
        self._Align = Alignment
        self._Border = Border
        self._Side = Side

        # Page setup
        ws.page_setup.paperSize = _PAPER_A4 if page_size == PageSize.A4 else _PAPER_A5
        ws.page_setup.orientation = (
            "landscape" if orientation == PageOrientation.LANDSCAPE else "portrait"
        )
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Page margins in inches (approx from mm)
        m = page_size.margin_mm / 25.4
        ws.page_margins.left = m
        ws.page_margins.right = m
        ws.page_margins.top = m
        ws.page_margins.bottom = m + 0.3  # extra for footer
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        # Sheet-level header/footer for printing
        ws.oddFooter.left.text = f"&8&B{_BRAND_NAME} · &I{_BRAND_TAGLINE}"
        ws.oddFooter.right.text = "&8Page &P of &N"
        ws.oddFooter.left.size = 8
        ws.oddFooter.right.size = 8

    # ── Public API ───────────────────────────────────────────────────────────────

    def write_document_header(
        self,
        company: CompanyHeaderData,
        title: str,
        *,
        generated_at: str | None = None,
    ) -> "ExcelSheetBuilder":
        """Write company name, title, and metadata rows at the top of the sheet."""
        from openpyxl.styles import PatternFill

        # Company name row
        cell = self._ws.cell(row=self._current_row, column=1, value=company.name)
        cell.font = self._Font(
            name="Calibri", size=14, bold=True, color=_C_BRAND_PRIMARY
        )
        self._ws.row_dimensions[self._current_row].height = 22
        self._current_row += 1

        # Legal name if different
        if company.legal_name and company.legal_name != company.name:
            lc = self._ws.cell(row=self._current_row, column=1, value=company.legal_name)
            lc.font = self._Font(name="Calibri", size=9, color=_C_MUTED)
            self._current_row += 1

        # Address
        for addr_line in company.address_block:
            ac = self._ws.cell(row=self._current_row, column=1, value=addr_line)
            ac.font = self._Font(name="Calibri", size=9)
            self._current_row += 1

        # Tax / reg
        meta: list[str] = []
        if company.tax_identifier:
            meta.append(f"Tax ID: {company.tax_identifier}")
        if company.registration_number:
            meta.append(f"Reg: {company.registration_number}")
        if meta:
            mc = self._ws.cell(row=self._current_row, column=1, value="  ·  ".join(meta))
            mc.font = self._Font(name="Calibri", size=9, color=_C_MUTED)
            self._current_row += 1

        self._current_row += 1  # blank separator

        # Document title
        tc = self._ws.cell(row=self._current_row, column=1, value=title)
        tc.font = self._Font(name="Calibri", size=13, bold=True, color=_C_BRAND_PRIMARY)
        self._ws.row_dimensions[self._current_row].height = 20
        self._current_row += 1

        # Generated at
        if generated_at:
            gc = self._ws.cell(
                row=self._current_row, column=1,
                value=f"Generated: {generated_at}",
            )
            gc.font = self._Font(name="Calibri", size=8, color=_C_MUTED, italic=True)
            self._current_row += 1

        self._current_row += 1  # blank row before table
        return self

    def write_table_header(self, columns: list[str]) -> "ExcelSheetBuilder":
        """Write the styled column header row for a data table."""
        self._num_columns = len(columns)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = self._ws.cell(row=self._current_row, column=col_idx, value=col_name)
            cell.font = self._Font(
                name="Calibri", size=9, bold=True, color=_C_WHITE
            )
            cell.fill = self._Fill(
                fill_type="solid", fgColor=_C_BRAND_PRIMARY
            )
            cell.alignment = self._Align(vertical="center", wrap_text=False)
        self._ws.row_dimensions[self._current_row].height = 18
        self._current_row += 1
        return self

    def write_table_row(
        self,
        values: list,
        *,
        numeric_columns: set[int] | None = None,
    ) -> "ExcelSheetBuilder":
        """Write a data row, alternating stripe colour."""
        num_cols = numeric_columns or set()
        is_even = (self._current_row % 2 == 0)
        bg = _C_WHITE if is_even else _C_STRIPE

        for col_idx, value in enumerate(values, start=1):
            cell = self._ws.cell(row=self._current_row, column=col_idx, value=value)
            cell.font = self._Font(name="Calibri", size=9)
            cell.fill = self._Fill(fill_type="solid", fgColor=bg)
            if col_idx - 1 in num_cols:
                cell.alignment = self._Align(horizontal="right")
        self._ws.row_dimensions[self._current_row].height = 16
        self._current_row += 1
        return self

    def write_totals_row(
        self,
        values: list,
        *,
        numeric_columns: set[int] | None = None,
    ) -> "ExcelSheetBuilder":
        """Write a styled totals row (bold, brand-tinted background, top border)."""
        from openpyxl.styles import Border, Side

        num_cols = numeric_columns or set()
        top_border = self._Border(
            top=self._Side(border_style="medium", color=_C_BRAND_PRIMARY)
        )
        for col_idx, value in enumerate(values, start=1):
            cell = self._ws.cell(row=self._current_row, column=col_idx, value=value)
            cell.font = self._Font(
                name="Calibri", size=9, bold=True, color=_C_BRAND_PRIMARY
            )
            cell.fill = self._Fill(fill_type="solid", fgColor=_C_BRAND_TOTAL)
            cell.border = top_border
            if col_idx - 1 in num_cols:
                cell.alignment = self._Align(horizontal="right")
        self._ws.row_dimensions[self._current_row].height = 17
        self._current_row += 1
        return self

    def write_blank_row(self, count: int = 1) -> "ExcelSheetBuilder":
        """Skip one or more rows."""
        self._current_row += count
        return self

    def write_key_value_pairs(
        self,
        pairs: list[tuple[str, str]],
    ) -> "ExcelSheetBuilder":
        """Write label-value pairs as metadata rows (used for document headers)."""
        for label, value in pairs:
            lc = self._ws.cell(row=self._current_row, column=1, value=label.upper())
            vc = self._ws.cell(row=self._current_row, column=2, value=value)
            lc.font = self._Font(name="Calibri", size=8, color=_C_MUTED)
            vc.font = self._Font(name="Calibri", size=9, bold=True)
            self._ws.row_dimensions[self._current_row].height = 15
            self._current_row += 1
        return self

    def write_branded_footer(self) -> "ExcelSheetBuilder":
        """Write Seeker Accounting branding rows at the bottom of the sheet."""
        self._current_row += 1  # blank separator
        cell = self._ws.cell(
            row=self._current_row,
            column=1,
            value=f"{_BRAND_NAME}  ·  {_BRAND_TAGLINE}",
        )
        cell.font = self._Font(
            name="Calibri", size=8, bold=False, color=_C_MUTED, italic=True
        )
        # Light top border
        from openpyxl.styles import Border, Side
        cell.border = self._Border(
            top=self._Side(border_style="thin", color=_C_BORDER)
        )
        self._current_row += 1
        return self

    def set_column_widths(self, widths: dict[int, float]) -> "ExcelSheetBuilder":
        """Set column widths (1-indexed, width in character units)."""
        from openpyxl.utils import get_column_letter

        for col_idx, width in widths.items():
            col_letter = get_column_letter(col_idx)
            self._ws.column_dimensions[col_letter].width = width
        return self

    def freeze_header_rows(self, rows: int = 1) -> "ExcelSheetBuilder":
        """Freeze the first N rows (column header freeze)."""
        from openpyxl.utils import get_column_letter

        self._ws.freeze_panes = f"A{rows + 1}"
        return self


class ExcelWorkbookBuilder:
    """Builds a multi-sheet Excel workbook."""

    def __init__(
        self,
        page_size: PageSize = PageSize.A4,
        orientation: PageOrientation = PageOrientation.PORTRAIT,
    ) -> None:
        from openpyxl import Workbook

        self._wb = Workbook()
        self._page_size = page_size
        self._orientation = orientation

        # Remove the default blank sheet
        if self._wb.active and self._wb.active.title == "Sheet":
            self._wb.remove(self._wb.active)

    def add_sheet(self, title: str) -> ExcelSheetBuilder:
        """Add a new worksheet and return a builder for it."""
        ws = self._wb.create_sheet(title=title)
        return ExcelSheetBuilder(ws, page_size=self._page_size, orientation=self._orientation)

    def save(self, output_path: str) -> None:
        """Save the workbook to disk."""
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        self._wb.save(output_path)
