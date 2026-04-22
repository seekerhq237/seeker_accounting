"""Excel (.xlsx) renderer for financial statements using openpyxl.

Fully independent from ``seeker_accounting.platform.printing``.
Each financial statement type is rendered with intentional, statement-specific
formatting through the section-based layout system.
"""

from __future__ import annotations

import datetime
from decimal import Decimal
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from seeker_accounting.modules.reporting.export.export_models import (
    StatementCompanyInfo,
    StatementExportRow,
    StatementPageOrientation,
    StatementPageSize,
    StatementTableSection,
)

_BRAND_PRIMARY_HEX = "1E3A5F"
_BRAND_LIGHT_HEX = "F0F3F7"
_SECTION_BG_HEX = "EDF1F6"
_SUBTOTAL_BG_HEX = "F5F6F8"
_TOTAL_BG_HEX = "E6EAF0"
_HIGHLIGHT_BG_HEX = "FFFCE8"
_WHITE_HEX = "FFFFFF"
_STRIPE_HEX = "F9FAFB"
_ZERO = Decimal("0.00")

# Reusable styles
_HEADER_FILL = PatternFill(start_color=_BRAND_PRIMARY_HEX, end_color=_BRAND_PRIMARY_HEX, fill_type="solid")
_HEADER_FONT = Font(bold=True, color=_WHITE_HEX, size=10)
_SECTION_FILL = PatternFill(start_color=_SECTION_BG_HEX, end_color=_SECTION_BG_HEX, fill_type="solid")
_SUBTOTAL_FILL = PatternFill(start_color=_SUBTOTAL_BG_HEX, end_color=_SUBTOTAL_BG_HEX, fill_type="solid")
_TOTAL_FILL = PatternFill(start_color=_TOTAL_BG_HEX, end_color=_TOTAL_BG_HEX, fill_type="solid")
_HIGHLIGHT_FILL = PatternFill(start_color=_HIGHLIGHT_BG_HEX, end_color=_HIGHLIGHT_BG_HEX, fill_type="solid")
_STRIPE_FILL = PatternFill(start_color=_STRIPE_HEX, end_color=_STRIPE_HEX, fill_type="solid")
_TOP_BORDER = Border(top=Side(style="medium", color=_BRAND_PRIMARY_HEX))
_DOUBLE_BORDER = Border(
    top=Side(style="double", color=_BRAND_PRIMARY_HEX),
    bottom=Side(style="double", color=_BRAND_PRIMARY_HEX),
)
_NUM_FORMAT = '#,##0.00;(#,##0.00);"-"'
_SECTION_HEADING_FILL = PatternFill(
    start_color=_BRAND_LIGHT_HEX, end_color=_BRAND_LIGHT_HEX, fill_type="solid",
)


class FinancialStatementExcelRenderer:
    """Renders financial statements to Excel workbooks.

    Accepts ``StatementTableSection`` objects so that each statement
    can define its own column structure and visual layout.
    """

    def render(
        self,
        *,
        title: str,
        subtitle: str | None,
        date_label: str,
        company: StatementCompanyInfo,
        sections: Sequence[StatementTableSection],
        summary_pairs: Sequence[tuple[str, str]],
        output_path: str,
        page_size: StatementPageSize,
        orientation: StatementPageOrientation,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        self._setup_page(ws, page_size, orientation)
        current_row = 1
        current_row = self._write_company_header(ws, company, current_row)
        current_row = self._write_title_block(ws, title, subtitle, date_label, current_row)

        # Track the widest column count across sections for summary/footer
        max_cols = max((len(s.column_headers) for s in sections), default=3)
        first_header_row: int | None = None

        for section in sections:
            col_count = len(section.column_headers)
            if section.heading:
                current_row = self._write_section_heading(ws, section.heading, col_count, current_row)
            header_row = current_row
            if first_header_row is None:
                first_header_row = header_row
            current_row = self._write_table_header(ws, section.column_headers, current_row)
            current_row = self._write_data_rows(ws, section.rows, col_count, current_row)
            self._set_column_widths(ws, col_count, section.column_widths)
            current_row += 1  # blank row between sections

        if summary_pairs:
            current_row = self._write_summary(ws, summary_pairs, max_cols, current_row)
        current_row = self._write_footer(ws, max_cols, current_row)

        # Freeze panes below first table header
        if first_header_row is not None:
            ws.freeze_panes = ws.cell(row=first_header_row + 1, column=1)

        wb.save(output_path)

    # ------------------------------------------------------------------
    # Page setup
    # ------------------------------------------------------------------

    def _setup_page(
        self,
        ws,
        page_size: StatementPageSize,
        orientation: StatementPageOrientation,
    ) -> None:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        if orientation == StatementPageOrientation.LANDSCAPE:
            ws.page_setup.orientation = "landscape"
        else:
            ws.page_setup.orientation = "portrait"

        margin_inches = page_size.margin_mm / 25.4
        ws.page_margins.left = margin_inches
        ws.page_margins.right = margin_inches
        ws.page_margins.top = margin_inches
        ws.page_margins.bottom = margin_inches

    # ------------------------------------------------------------------
    # Company header
    # ------------------------------------------------------------------

    def _write_company_header(self, ws, company: StatementCompanyInfo, row: int) -> int:
        cell = ws.cell(row=row, column=1, value=company.name)
        cell.font = Font(bold=True, size=14, color=_BRAND_PRIMARY_HEX)
        row += 1

        if company.legal_name and company.legal_name != company.name:
            cell = ws.cell(row=row, column=1, value=company.legal_name)
            cell.font = Font(size=10, color="888888")
            row += 1

        for line in company.address_block:
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = Font(size=9, color="666666")
            row += 1

        id_parts: list[str] = []
        if company.tax_identifier:
            id_parts.append(f"Tax ID: {company.tax_identifier}")
        if company.registration_number:
            id_parts.append(f"Reg: {company.registration_number}")
        if id_parts:
            cell = ws.cell(row=row, column=1, value=" · ".join(id_parts))
            cell.font = Font(size=9, color="666666")
            row += 1

        row += 1  # blank row
        return row

    # ------------------------------------------------------------------
    # Title block
    # ------------------------------------------------------------------

    def _write_title_block(
        self,
        ws,
        title: str,
        subtitle: str | None,
        date_label: str,
        row: int,
    ) -> int:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=13, color=_BRAND_PRIMARY_HEX)
        row += 1

        if subtitle:
            cell = ws.cell(row=row, column=1, value=subtitle)
            cell.font = Font(size=10, color="555555")
            row += 1

        cell = ws.cell(row=row, column=1, value=date_label)
        cell.font = Font(size=9, color="666666")
        row += 1

        row += 1  # blank row
        return row

    # ------------------------------------------------------------------
    # Section heading
    # ------------------------------------------------------------------

    def _write_section_heading(self, ws, heading: str, col_count: int, row: int) -> int:
        cell = ws.cell(row=row, column=1, value=heading)
        cell.font = Font(bold=True, size=11, color=_BRAND_PRIMARY_HEX)
        cell.fill = _SECTION_HEADING_FILL
        # Apply fill across all columns
        for c in range(2, col_count + 1):
            ws.cell(row=row, column=c).fill = _SECTION_HEADING_FILL
        return row + 1

    # ------------------------------------------------------------------
    # Table header
    # ------------------------------------------------------------------

    def _write_table_header(self, ws, column_headers: Sequence[str], row: int) -> int:
        for i, hdr in enumerate(column_headers, start=1):
            cell = ws.cell(row=row, column=i, value=hdr)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            if i >= 3:
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
        return row + 1

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------

    def _write_data_rows(
        self,
        ws,
        rows: Sequence[StatementExportRow],
        col_count: int,
        start_row: int,
    ) -> int:
        row = start_row
        stripe_counter = 0

        for data in rows:
            kind = data.row_kind
            is_section = kind == "section"
            is_total = kind == "total"
            is_bold = kind in ("section", "group", "formula", "subtotal", "subsection", "total")
            is_formula = kind in ("formula", "subtotal")

            # Ref
            ref_text = data.ref_code if not is_section else ""
            ref_cell = ws.cell(row=row, column=1, value=ref_text)
            ref_cell.font = Font(size=9, color="888888")

            # Label with indent
            indent_str = "    " * data.indent_level
            lbl_cell = ws.cell(row=row, column=2, value=indent_str + data.label)
            lbl_cell.font = Font(bold=is_bold, size=10)

            # Amounts
            for i in range(2, col_count):
                amt_idx = i - 2
                val = data.amounts[amt_idx] if amt_idx < len(data.amounts) else None
                amt_cell = ws.cell(row=row, column=i + 1)
                if val is not None:
                    amt_cell.value = float(val)
                    amt_cell.number_format = _NUM_FORMAT
                amt_cell.alignment = Alignment(horizontal="right")
                amt_cell.font = Font(bold=is_bold, size=10)

            # Row styling
            fill = None
            if is_total:
                fill = _TOTAL_FILL
            elif is_section:
                fill = _SECTION_FILL
                stripe_counter = 0
            elif is_formula:
                fill = _SUBTOTAL_FILL
            elif data.is_highlight:
                fill = _HIGHLIGHT_FILL
            else:
                stripe_counter += 1
                if stripe_counter % 2 == 0:
                    fill = _STRIPE_FILL

            if fill:
                for c in range(1, col_count + 1):
                    ws.cell(row=row, column=c).fill = fill

            # Borders
            if is_total:
                for c in range(1, col_count + 1):
                    ws.cell(row=row, column=c).border = _DOUBLE_BORDER
            elif is_formula:
                for c in range(1, col_count + 1):
                    ws.cell(row=row, column=c).border = _TOP_BORDER

            row += 1
        return row

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------

    def _write_summary(
        self,
        ws,
        pairs: Sequence[tuple[str, str]],
        col_count: int,
        row: int,
    ) -> int:
        row += 1  # blank row

        for idx, (label, value) in enumerate(pairs):
            is_last = idx == len(pairs) - 1
            lbl_cell = ws.cell(row=row, column=1, value=label)
            lbl_cell.font = Font(bold=True, size=10)

            val_cell = ws.cell(row=row, column=col_count, value=value)
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.font = Font(bold=is_last, size=10)

            if is_last:
                for c in range(1, col_count + 1):
                    ws.cell(row=row, column=c).fill = PatternFill(
                        start_color=_BRAND_LIGHT_HEX,
                        end_color=_BRAND_LIGHT_HEX,
                        fill_type="solid",
                    )
            row += 1
        return row

    # ------------------------------------------------------------------
    # Footer
    # ------------------------------------------------------------------

    def _write_footer(self, ws, col_count: int, row: int) -> int:
        row += 1
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        cell = ws.cell(row=row, column=1, value=f"Seeker Accounting · Generated {now}")
        cell.font = Font(size=8, color="999999")
        return row + 1

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------

    def _set_column_widths(
        self,
        ws,
        col_count: int,
        custom_widths: Sequence[float] | tuple[float, ...] | None,
    ) -> None:
        if custom_widths:
            for i, w in enumerate(custom_widths, start=1):
                cur = ws.column_dimensions[get_column_letter(i)].width or 0
                ws.column_dimensions[get_column_letter(i)].width = max(cur, w)
        else:
            defaults = [10.0, 50.0] + [18.0] * max(0, col_count - 2)
            for i, w in enumerate(defaults[:col_count], start=1):
                cur = ws.column_dimensions[get_column_letter(i)].width or 0
                ws.column_dimensions[get_column_letter(i)].width = max(cur, w)
