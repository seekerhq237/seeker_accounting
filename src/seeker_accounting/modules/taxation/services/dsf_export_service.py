"""DSF (Déclaration Statistique et Fiscale) export service.

Phase-1 export produces a multi-sheet Excel workbook from the
company tax profile and the year's posted VAT data. The DGI's own
DSF Excel template is a separate, regulator-controlled artifact;
this service produces an *internal* DSF working file that captures
the same source facts in a stable, reviewable layout. A future
Phase-4 slice will adapt this output into a DGI-shaped template.

The export reads only from:

* ``company_tax_profiles`` (tax identity)
* ``tax_obligations``     (compliance calendar)
* ``tax_returns``         (filings)
* ``tax_return_lines``    (statutory boxes)
* ``tax_payments``        (settlements)

It does **not** rebuild VAT box totals from posted documents — that
is the job of ``TaxReturnService.draft_vat_return``. The DSF export
simply renders what has already been drafted/filed.
"""

from __future__ import annotations

import os
from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING, Callable

from sqlalchemy.orm import Session

from seeker_accounting.app.context.app_context import AppContext
from seeker_accounting.db.unit_of_work import UnitOfWorkFactory
from seeker_accounting.modules.administration.services.permission_service import (
    PermissionService,
)
from seeker_accounting.modules.companies.repositories.company_repository import (
    CompanyRepository,
)
from seeker_accounting.modules.taxation.constants import (
    DSF_FORM_LIBERATORY,
    DSF_FORM_NONE,
    DSF_FORM_REAL,
    DSF_FORM_SIMPLIFIED,
    RETURN_STATUS_CANCELLED,
    RETURN_STATUS_FILED,
    TAX_TYPE_CIT_BALANCE,
    TAX_TYPE_CIT_INSTALLMENT,
    TAX_TYPE_VAT,
)
from seeker_accounting.modules.taxation.dto.dsf_export_dto import (
    DSFExportResultDTO,
    DSFReadinessIssue,
    GenerateDSFExportCommand,
)
from seeker_accounting.modules.taxation.repositories.company_tax_profile_repository import (
    CompanyTaxProfileRepository,
)
from seeker_accounting.modules.taxation.repositories.tax_obligation_repository import (
    TaxObligationRepository,
)
from seeker_accounting.modules.taxation.repositories.tax_payment_repository import (
    TaxPaymentRepository,
)
from seeker_accounting.modules.taxation.repositories.tax_return_repository import (
    TaxReturnRepository,
)
from seeker_accounting.modules.taxation.repositories.withholding_tax_certificate_repository import (
    WithholdingTaxCertificateRepository,
)
from seeker_accounting.modules.reporting.dto.reporting_filter_dto import (
    ReportingFilterDTO,
)
from seeker_accounting.platform.exceptions import (
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)

if TYPE_CHECKING:
    from seeker_accounting.modules.audit.services.audit_service import AuditService
    from seeker_accounting.modules.reporting.services.ohada_balance_sheet_service import (
        OhadaBalanceSheetService,
    )
    from seeker_accounting.modules.reporting.services.ohada_income_statement_service import (
        OhadaIncomeStatementService,
    )


CompanyTaxProfileRepositoryFactory = Callable[[Session], CompanyTaxProfileRepository]
TaxObligationRepositoryFactory = Callable[[Session], TaxObligationRepository]
TaxReturnRepositoryFactory = Callable[[Session], TaxReturnRepository]
TaxPaymentRepositoryFactory = Callable[[Session], TaxPaymentRepository]
CompanyRepositoryFactory = Callable[[Session], CompanyRepository]
WithholdingTaxCertificateRepositoryFactory = Callable[
    [Session], WithholdingTaxCertificateRepository
]


_ZERO = Decimal("0.00")


class DSFExportService:
    PERMISSION_EXPORT = "taxation.dsf.export"

    def __init__(
        self,
        unit_of_work_factory: UnitOfWorkFactory,
        app_context: AppContext,
        company_tax_profile_repository_factory: CompanyTaxProfileRepositoryFactory,
        tax_obligation_repository_factory: TaxObligationRepositoryFactory,
        tax_return_repository_factory: TaxReturnRepositoryFactory,
        tax_payment_repository_factory: TaxPaymentRepositoryFactory,
        company_repository_factory: CompanyRepositoryFactory,
        permission_service: PermissionService,
        audit_service: "AuditService | None" = None,
        ohada_balance_sheet_service: "OhadaBalanceSheetService | None" = None,
        ohada_income_statement_service: "OhadaIncomeStatementService | None" = None,
        withholding_tax_certificate_repository_factory: (
            WithholdingTaxCertificateRepositoryFactory | None
        ) = None,
    ) -> None:
        self._unit_of_work_factory = unit_of_work_factory
        self._app_context = app_context
        self._company_tax_profile_repository_factory = (
            company_tax_profile_repository_factory
        )
        self._tax_obligation_repository_factory = tax_obligation_repository_factory
        self._tax_return_repository_factory = tax_return_repository_factory
        self._tax_payment_repository_factory = tax_payment_repository_factory
        self._company_repository_factory = company_repository_factory
        self._permission_service = permission_service
        self._audit_service = audit_service
        self._ohada_balance_sheet_service = ohada_balance_sheet_service
        self._ohada_income_statement_service = ohada_income_statement_service
        self._withholding_tax_certificate_repository_factory = (
            withholding_tax_certificate_repository_factory
        )

    # ------------------------------------------------------------------
    # Entry points
    # ------------------------------------------------------------------

    def check_readiness(
        self, company_id: int, fiscal_year: int
    ) -> tuple[DSFReadinessIssue, ...]:
        """Run pre-export validation without writing a file."""
        self._permission_service.require_permission(self.PERMISSION_EXPORT)
        self._validate_year(fiscal_year)
        with self._unit_of_work_factory() as uow:
            self._require_company_exists(uow.session, company_id)
            return self._collect_readiness_issues(uow.session, company_id, fiscal_year)

    def generate(
        self,
        company_id: int,
        command: GenerateDSFExportCommand,
        actor_user_id: int | None = None,
    ) -> DSFExportResultDTO:
        self._permission_service.require_permission(self.PERMISSION_EXPORT)
        self._validate_year(command.fiscal_year)

        output_path = (command.output_path or "").strip()
        if not output_path:
            raise ValidationError("Output path is required.")
        if not output_path.lower().endswith(".xlsx"):
            raise ValidationError("DSF export output path must end with .xlsx.")

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        with self._unit_of_work_factory() as uow:
            self._require_company_exists(uow.session, company_id)

            profile_repo = self._company_tax_profile_repository_factory(uow.session)
            obligation_repo = self._tax_obligation_repository_factory(uow.session)
            return_repo = self._tax_return_repository_factory(uow.session)
            payment_repo = self._tax_payment_repository_factory(uow.session)

            profile = profile_repo.get_by_company(company_id)
            obligations, returns, payments = self._load_year_data(
                uow.session, company_id, command.fiscal_year
            )
            wht_certificates = self._load_withholding_certificates(
                uow.session, company_id, command.fiscal_year
            )

            issues = self._collect_readiness_issues(
                uow.session,
                company_id,
                command.fiscal_year,
                profile=profile,
                obligations=obligations,
                returns=returns,
            )
            has_blocking = any(i.severity == "error" for i in issues)

            (
                balance_sheet_amounts,
                income_statement_amounts,
                fiche_issues,
                balance_sheet_full_dto,
                income_statement_full_dto,
            ) = self._compute_account_balances(company_id, command.fiscal_year)
            if fiche_issues:
                issues = (*issues, *fiche_issues)

            sheets = self._write_workbook(
                output_path=output_path,
                fiscal_year=command.fiscal_year,
                profile=profile,
                obligations=obligations,
                returns=returns,
                payments=payments,
                issues=issues,
                balance_sheet_amounts=balance_sheet_amounts,
                income_statement_amounts=income_statement_amounts,
                wht_certificates=wht_certificates,
                balance_sheet_full_dto=balance_sheet_full_dto,
                income_statement_full_dto=income_statement_full_dto,
            )

            dsf_form_applied = self._effective_dsf_form(profile)
            tax_regime_applied = (
                getattr(profile, "tax_regime_code", None) if profile else None
            )

            self._record_audit(
                company_id,
                f"Generated DSF export for {command.fiscal_year} "
                f"({len(returns)} returns, {len(payments)} payments, "
                f"form={dsf_form_applied or 'NONE'}).",
            )

            return DSFExportResultDTO(
                company_id=company_id,
                fiscal_year=command.fiscal_year,
                output_path=output_path,
                sheets_written=sheets,
                obligation_count=len(obligations),
                return_count=len(returns),
                payment_count=len(payments),
                readiness_issues=tuple(issues),
                has_blocking_issues=has_blocking,
                dsf_form_applied=dsf_form_applied,
                tax_regime_applied=tax_regime_applied,
                withholding_certificate_count=len(wht_certificates),
            )

    # ------------------------------------------------------------------
    # Data loading
    # ------------------------------------------------------------------

    def _load_year_data(
        self, session: Session, company_id: int, fiscal_year: int
    ) -> tuple[list, list, list]:
        period_start = date(fiscal_year, 1, 1)
        period_end = date(fiscal_year, 12, 31)

        obligation_repo = self._tax_obligation_repository_factory(session)
        return_repo = self._tax_return_repository_factory(session)
        payment_repo = self._tax_payment_repository_factory(session)

        all_obligations = obligation_repo.list_by_company(company_id)
        obligations = [
            o
            for o in all_obligations
            if o.period_start >= period_start and o.period_end <= period_end
        ]

        all_returns = return_repo.list_by_company(company_id)
        returns = [
            r
            for r in all_returns
            if r.period_start >= period_start and r.period_end <= period_end
        ]
        # Pre-load lines for each return.
        returns = [return_repo.get_by_id(company_id, r.id) for r in returns]
        returns = [r for r in returns if r is not None]

        all_payments = payment_repo.list_by_company(company_id)
        payments = [
            p
            for p in all_payments
            if period_start <= p.payment_date <= period_end
        ]
        return obligations, returns, payments

    def _load_withholding_certificates(
        self, session: Session, company_id: int, fiscal_year: int
    ) -> list:
        """Load WHT certificate rows for the fiscal year, when wired.

        Returns ``[]`` when the optional repository factory is not
        provided (older wiring or test stubs). The certificate sheet
        is then simply omitted from the workbook.
        """
        if self._withholding_tax_certificate_repository_factory is None:
            return []
        period_start = date(fiscal_year, 1, 1)
        period_end = date(fiscal_year, 12, 31)
        repo = self._withholding_tax_certificate_repository_factory(session)
        return repo.list_by_company(
            company_id,
            date_from=period_start,
            date_to=period_end,
        )

    # ------------------------------------------------------------------
    # Readiness
    # ------------------------------------------------------------------

    def _collect_readiness_issues(
        self,
        session: Session,
        company_id: int,
        fiscal_year: int,
        *,
        profile=None,
        obligations=None,
        returns=None,
    ) -> tuple[DSFReadinessIssue, ...]:
        issues: list[DSFReadinessIssue] = []

        if profile is None:
            profile_repo = self._company_tax_profile_repository_factory(session)
            profile = profile_repo.get_by_company(company_id)

        if obligations is None or returns is None:
            obligations, returns, _ = self._load_year_data(
                session, company_id, fiscal_year
            )

        if profile is None:
            issues.append(
                DSFReadinessIssue(
                    severity="error",
                    code="NO_TAX_PROFILE",
                    message=(
                        "Company has no tax profile. Configure NIU, regime, and DSF "
                        "form selection before exporting."
                    ),
                )
            )
        else:
            if not (profile.niu or "").strip():
                issues.append(
                    DSFReadinessIssue(
                        severity="error",
                        code="MISSING_NIU",
                        message="Company NIU is required for DSF filing.",
                    )
                )
            if profile.dsf_form_code is None:
                issues.append(
                    DSFReadinessIssue(
                        severity="warning",
                        code="MISSING_DSF_FORM",
                        message="DSF form family is not set on the tax profile.",
                    )
                )
            if profile.tax_regime_code is None:
                issues.append(
                    DSFReadinessIssue(
                        severity="warning",
                        code="MISSING_REGIME",
                        message="Tax regime is not set on the tax profile.",
                    )
                )

        # VAT obligation coverage
        vat_obligations = [o for o in obligations if o.tax_type_code == TAX_TYPE_VAT]
        if profile is not None and bool(getattr(profile, "is_vat_liable", False)):
            if len(vat_obligations) < 12:
                issues.append(
                    DSFReadinessIssue(
                        severity="warning",
                        code="INCOMPLETE_VAT_OBLIGATIONS",
                        message=(
                            f"Only {len(vat_obligations)} of 12 monthly VAT "
                            "obligations exist for the year."
                        ),
                    )
                )

        # Returns: count drafts vs filed
        draft_count = sum(
            1
            for r in returns
            if r.status_code != RETURN_STATUS_FILED
            and r.status_code != RETURN_STATUS_CANCELLED
        )
        if draft_count > 0:
            issues.append(
                DSFReadinessIssue(
                    severity="warning",
                    code="UNFILED_RETURNS",
                    message=(
                        f"{draft_count} return(s) are still in draft for the year."
                    ),
                )
            )

        return tuple(issues)

    # ------------------------------------------------------------------
    # Workbook generation
    # ------------------------------------------------------------------

    def _compute_account_balances(
        self,
        company_id: int,
        fiscal_year: int,
    ) -> tuple[
        dict[str, dict[str, Decimal]] | None,
        dict[str, Decimal] | None,
        tuple[DSFReadinessIssue, ...],
        object | None,  # full balance sheet DTO (T25)
        object | None,  # full income statement DTO (T25)
    ]:
        """Aggregate posted GL balances for Fiche R3 (balance sheet) and
        Fiche R4 (income statement).

        Returns a triple ``(bs_amounts, is_amounts, extra_issues)``:

        * ``bs_amounts`` — ``{ref_code: {"gross", "contra", "net"}}`` or
          ``None`` when the OHADA balance sheet service is not wired or
          its computation fails.
        * ``is_amounts`` — ``{ref_code: signed_amount}`` or ``None``
          under the same conditions.
        * ``extra_issues`` — readiness warnings to surface on the
          Readiness sheet when amounts could not be populated.

        Failures are converted into ``warning`` readiness issues so the
        export still succeeds with blank cells; the DSF export must not
        be blocked because a downstream report cannot be assembled.
        """
        extra: list[DSFReadinessIssue] = []
        bs_amounts: dict[str, dict[str, Decimal]] | None = None
        is_amounts: dict[str, Decimal] | None = None
        bs_full_dto = None  # T25
        is_full_dto = None  # T25

        period_start = date(fiscal_year, 1, 1)
        period_end = date(fiscal_year, 12, 31)
        filter_dto = ReportingFilterDTO(
            company_id=company_id,
            date_from=period_start,
            date_to=period_end,
            posted_only=True,
        )

        if self._ohada_balance_sheet_service is not None:
            try:
                bs_dto = self._ohada_balance_sheet_service.get_statement(filter_dto)
                bs_full_dto = bs_dto  # T25 — full DTO for standalone sheet
                bs_amounts = {}
                for line in (*bs_dto.asset_lines, *bs_dto.liability_lines):
                    ref = line.reference_code
                    if not ref:
                        continue
                    bs_amounts[ref] = {
                        "gross": line.gross_amount or _ZERO,
                        "contra": line.contra_amount or _ZERO,
                        "net": line.net_amount or _ZERO,
                    }
            except (PermissionDeniedError, ValidationError, NotFoundError) as exc:
                bs_amounts = None
                extra.append(
                    DSFReadinessIssue(
                        severity="warning",
                        code="FICHE_R3_AMOUNTS_UNAVAILABLE",
                        message=(
                            "Fiche R3 balance sheet amounts could not be "
                            f"computed and were left blank: {exc}"
                        ),
                    )
                )

        if self._ohada_income_statement_service is not None:
            try:
                is_dto = self._ohada_income_statement_service.get_statement(filter_dto)
                is_full_dto = is_dto  # T25 — full DTO for standalone sheet
                is_amounts = {line.code: line.signed_amount for line in is_dto.lines}
            except (PermissionDeniedError, ValidationError, NotFoundError) as exc:
                is_amounts = None
                extra.append(
                    DSFReadinessIssue(
                        severity="warning",
                        code="FICHE_R4_AMOUNTS_UNAVAILABLE",
                        message=(
                            "Fiche R4 income statement amounts could not be "
                            f"computed and were left blank: {exc}"
                        ),
                    )
                )

        return bs_amounts, is_amounts, tuple(extra), bs_full_dto, is_full_dto

    def _write_workbook(
        self,
        *,
        output_path: str,
        fiscal_year: int,
        profile,
        obligations: list,
        returns: list,
        payments: list,
        issues: tuple[DSFReadinessIssue, ...],
        balance_sheet_amounts: dict[str, dict[str, Decimal]] | None = None,
        income_statement_amounts: dict[str, Decimal] | None = None,
        wht_certificates: list | None = None,
        balance_sheet_full_dto=None,  # T25
        income_statement_full_dto=None,  # T25
    ) -> tuple[str, ...]:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )

        def _write_header(ws, headers: list[str]) -> None:
            for col, value in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=value)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"

        sheets_written: list[str] = []

        # ── Company profile ──
        ws = wb.create_sheet("Company Profile")
        ws.append(["Field", "Value"])
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        rows = self._profile_rows(profile, fiscal_year)
        for label, value in rows:
            ws.append([label, value])
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 40
        sheets_written.append("Company Profile")

        # ── VAT Summary ──
        ws = wb.create_sheet("VAT Summary")
        _write_header(
            ws,
            [
                "Period start",
                "Period end",
                "Due date",
                "Obligation status",
                "Return status",
                "Total due",
                "Total paid",
                "Filed at",
                "OTP reference",
                "External reference",
            ],
        )
        returns_by_obligation = {r.obligation_id: r for r in returns}
        for o in sorted(obligations, key=lambda x: x.period_start):
            if o.tax_type_code != TAX_TYPE_VAT:
                continue
            r = returns_by_obligation.get(o.id)
            ws.append(
                [
                    o.period_start,
                    o.period_end,
                    o.due_date,
                    o.status_code,
                    r.status_code if r else "",
                    float(r.total_due_amount) if r else 0.0,
                    float(r.total_paid_amount) if r else 0.0,
                    r.filed_at if r and r.filed_at else "",
                    r.otp_reference if r and r.otp_reference else "",
                    r.external_reference if r and r.external_reference else "",
                ]
            )
        for col in ("A", "B", "C", "H"):
            ws.column_dimensions[col].width = 16
        for col in ("D", "E", "I", "J"):
            ws.column_dimensions[col].width = 18
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        sheets_written.append("VAT Summary")

        # ── VAT Detail ──
        ws = wb.create_sheet("VAT Detail")
        _write_header(
            ws,
            [
                "Period start",
                "Period end",
                "Return status",
                "Box code",
                "Box label",
                "Amount",
            ],
        )
        for r in sorted(returns, key=lambda x: x.period_start):
            for line in sorted(r.lines, key=lambda x: x.sort_order):
                ws.append(
                    [
                        r.period_start,
                        r.period_end,
                        r.status_code,
                        line.box_code,
                        line.label,
                        float(line.amount),
                    ]
                )
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 24
        ws.column_dimensions["E"].width = 36
        ws.column_dimensions["F"].width = 16
        sheets_written.append("VAT Detail")

        # ── Payments ──
        ws = wb.create_sheet("Payments")
        _write_header(
            ws,
            [
                "Payment date",
                "Amount",
                "Method",
                "Reference",
                "Tax return id",
                "Notes",
            ],
        )
        for p in sorted(payments, key=lambda x: x.payment_date):
            ws.append(
                [
                    p.payment_date,
                    float(p.amount),
                    p.payment_method_code,
                    p.reference or "",
                    p.tax_return_id or "",
                    p.notes or "",
                ]
            )
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 36
        sheets_written.append("Payments")

        # ── Withholding Certificates (Slice T13) ──
        # Only emit the sheet when the WHT repository factory was
        # wired into this service. When unwired (older wiring or test
        # stubs) ``wht_certificates`` is ``None``/empty and the sheet
        # is suppressed entirely so historical exports remain
        # byte-identical.
        if wht_certificates:
            ws = wb.create_sheet("Withholding Certificates")
            _write_header(
                ws,
                [
                    "Direction",
                    "Date",
                    "Number",
                    "Counterparty kind",
                    "Counterparty name",
                    "NIU",
                    "Tax code id",
                    "Status",
                    "Taxable base",
                    "Tax amount",
                    "Notes",
                ],
            )
            for c in sorted(
                wht_certificates,
                key=lambda x: (x.direction, x.certificate_date, x.id),
            ):
                ws.append(
                    [
                        c.direction,
                        c.certificate_date,
                        c.certificate_number,
                        c.counterparty_kind,
                        c.counterparty_name,
                        c.counterparty_niu or "",
                        c.tax_code_id,
                        c.status_code,
                        float(c.taxable_base),
                        float(c.tax_amount),
                        (c.notes or ""),
                    ]
                )
            for col in ("A", "B", "D", "G", "H"):
                ws.column_dimensions[col].width = 14
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["E"].width = 32
            ws.column_dimensions["F"].width = 18
            ws.column_dimensions["I"].width = 16
            ws.column_dimensions["J"].width = 16
            ws.column_dimensions["K"].width = 36
            sheets_written.append("Withholding Certificates")

        # ── Income Statement (P&L) — T25 ──
        # Standalone P&L sheet sourced from the OHADA income statement
        # service. Emitted only when the upstream service is wired and
        # produced a non-empty DTO.
        if income_statement_full_dto is not None and getattr(
            income_statement_full_dto, "lines", ()
        ):
            ws = wb.create_sheet("Income Statement")
            _write_header(
                ws,
                ["Code", "Section", "Label", "Signed amount", "Drilldown"],
            )
            for line in income_statement_full_dto.lines:
                ws.append(
                    [
                        getattr(line, "code", "") or "",
                        getattr(line, "section_title", None)
                        or getattr(line, "section_code", "")
                        or "",
                        getattr(line, "label", "") or "",
                        float(getattr(line, "signed_amount", None) or 0),
                        "Yes" if getattr(line, "can_drilldown", False) else "",
                    ]
                )
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 32
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 12
            sheets_written.append("Income Statement")

        # ── Balance Sheet — T25 ──
        # Standalone BS sheet sourced from the OHADA balance sheet
        # service. Asset and liability/equity lines are written in DGI
        # display order, separated by a blank divider row for visual
        # clarity.
        if balance_sheet_full_dto is not None and (
            getattr(balance_sheet_full_dto, "asset_lines", ())
            or getattr(balance_sheet_full_dto, "liability_lines", ())
        ):
            ws = wb.create_sheet("Balance Sheet")
            _write_header(
                ws,
                ["Side", "Code", "Ref", "Label", "Gross", "Contra", "Net"],
            )

            def _write_bs_lines(side: str, lines) -> None:
                for line in lines:
                    ws.append(
                        [
                            side,
                            getattr(line, "code", "") or "",
                            getattr(line, "reference_code", "") or "",
                            getattr(line, "label", "") or "",
                            float(getattr(line, "gross_amount", None) or 0),
                            float(getattr(line, "contra_amount", None) or 0),
                            float(getattr(line, "net_amount", None) or 0),
                        ]
                    )

            _write_bs_lines("ASSET", balance_sheet_full_dto.asset_lines)
            ws.append([""] * 7)  # divider
            _write_bs_lines("LIAB/EQ", balance_sheet_full_dto.liability_lines)
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 60
            ws.column_dimensions["E"].width = 16
            ws.column_dimensions["F"].width = 16
            ws.column_dimensions["G"].width = 16
            sheets_written.append("Balance Sheet")

        # ── Readiness ──
        ws = wb.create_sheet("Readiness")
        _write_header(ws, ["Severity", "Code", "Message"])
        if not issues:
            ws.append(["info", "READY", "No readiness issues detected."])
        else:
            for issue in issues:
                ws.append([issue.severity, issue.code, issue.message])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 28
        ws.column_dimensions["C"].width = 80
        sheets_written.append("Readiness")

        # ── Form-family-specific fiches (Phase 4) ──
        # The base sheets above are present for every export. The fiches
        # below adapt to the company's DSF form selection so reviewers
        # see the exact DGI structure (R1 / R2 / R3 + annexes) for their
        # regime. Sheets are added in DGI order to make navigation
        # predictable.
        form_code = self._effective_dsf_form(profile)
        family_sheets = self._write_form_family_fiches(
            wb=wb,
            header_font=header_font,
            header_fill=header_fill,
            write_header=_write_header,
            form_code=form_code,
            fiscal_year=fiscal_year,
            profile=profile,
            obligations=obligations,
            returns=returns,
            payments=payments,
            balance_sheet_amounts=balance_sheet_amounts,
            income_statement_amounts=income_statement_amounts,
        )
        sheets_written.extend(family_sheets)

        wb.save(output_path)
        return tuple(sheets_written)

    @staticmethod
    def _profile_rows(profile, fiscal_year: int) -> list[tuple[str, str]]:
        def _v(value) -> str:
            if value is None:
                return ""
            if isinstance(value, bool):
                return "Yes" if value else "No"
            return str(value)

        if profile is None:
            return [
                ("Fiscal year", str(fiscal_year)),
                ("Tax profile", "(not configured)"),
            ]
        return [
            ("Fiscal year", str(fiscal_year)),
            ("NIU", _v(profile.niu)),
            ("Tax center code", _v(profile.tax_center_code)),
            ("Taxpayer segment", _v(profile.taxpayer_segment_code)),
            ("Tax regime", _v(profile.tax_regime_code)),
            ("VAT liable", _v(profile.is_vat_liable)),
            ("VAT effective from", _v(profile.vat_effective_from)),
            ("CIT rate profile", _v(profile.cit_rate_profile_code)),
            (
                "CIT installment profile",
                _v(profile.cit_installment_profile_code),
            ),
            ("SME qualified", _v(profile.sme_qualified_flag)),
            ("DSF form", _v(profile.dsf_form_code)),
            ("DSF submission mode", _v(profile.dsf_submission_mode_code)),
            ("OTP enabled", _v(profile.otp_enabled_flag)),
            (
                "Default withholding applicable",
                _v(profile.default_withholding_applicable_flag),
            ),
        ]

    # ------------------------------------------------------------------
    # Form-family fiches (Phase 4 — regime-conditional layouts)
    # ------------------------------------------------------------------

    @staticmethod
    def _effective_dsf_form(profile) -> str | None:
        """Return the DSF form code that should drive fiche generation.

        ``None`` means *no fiches* — only the base sheets are written.
        ``DSF_FORM_NONE`` from the profile is treated the same as
        unset, which matches DGI's "no DSF required" path (e.g. tiny
        liberatory taxpayers below the IGS threshold). Unknown / legacy
        codes are also treated as "no fiches" — fiches are only emitted
        for the three canonical DGI form families.
        """
        if profile is None:
            return None
        code = getattr(profile, "dsf_form_code", None)
        if not code or code == DSF_FORM_NONE:
            return None
        if code in (DSF_FORM_REAL, DSF_FORM_SIMPLIFIED, DSF_FORM_LIBERATORY):
            return code
        return None

    def _write_form_family_fiches(
        self,
        *,
        wb,
        header_font,
        header_fill,
        write_header,
        form_code: str | None,
        fiscal_year: int,
        profile,
        obligations: list,
        returns: list,
        payments: list,
        balance_sheet_amounts: dict[str, dict[str, Decimal]] | None = None,
        income_statement_amounts: dict[str, Decimal] | None = None,
    ) -> list[str]:
        """Append regime-specific fiches to the workbook.

        Returns the list of sheet names that were added (in order). The
        DGI fiche structure is documented in the
        ``DSF 2025 Guide`` (taxpayer guide on e-filing of the DSF) —
        every variant carries Fiche R1 (identity); REAL/SIMPLIFIED add
        Fiche R2 (revenue & activity codes) and Fiche R3 (balance
        sheet); REAL additionally carries the full income statement and
        CIT computation summary; LIBERATORY collapses everything into
        an IGS / patente summary.
        """
        if form_code is None:
            return []

        sheets: list[str] = []

        # ── Fiche R1 — Identity (all variants) ──
        sheets.append(
            self._write_fiche_r1_identity(
                wb=wb,
                header_font=header_font,
                header_fill=header_fill,
                form_code=form_code,
                fiscal_year=fiscal_year,
                profile=profile,
            )
        )

        if form_code == DSF_FORM_REAL:
            sheets.append(
                self._write_fiche_r2_revenue(
                    wb=wb,
                    write_header=write_header,
                    profile=profile,
                    obligations=obligations,
                    returns=returns,
                    fiscal_year=fiscal_year,
                )
            )
            sheets.append(
                self._write_fiche_r3_balance_sheet(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    amounts=balance_sheet_amounts,
                )
            )
            sheets.append(
                self._write_fiche_income_statement(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    simplified=False,
                    amounts=income_statement_amounts,
                )
            )
            sheets.append(
                self._write_fiche_cit_summary(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    obligations=obligations,
                    returns=returns,
                    payments=payments,
                )
            )

        elif form_code == DSF_FORM_SIMPLIFIED:
            sheets.append(
                self._write_fiche_r2_revenue(
                    wb=wb,
                    write_header=write_header,
                    profile=profile,
                    obligations=obligations,
                    returns=returns,
                    fiscal_year=fiscal_year,
                )
            )
            sheets.append(
                self._write_fiche_income_statement(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    simplified=True,
                    amounts=income_statement_amounts,
                )
            )
            sheets.append(
                self._write_fiche_cit_summary(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    obligations=obligations,
                    returns=returns,
                    payments=payments,
                )
            )

        elif form_code == DSF_FORM_LIBERATORY:
            sheets.append(
                self._write_fiche_liberatory_summary(
                    wb=wb,
                    write_header=write_header,
                    fiscal_year=fiscal_year,
                    profile=profile,
                    obligations=obligations,
                    returns=returns,
                    payments=payments,
                )
            )

        return sheets

    @staticmethod
    def _write_fiche_r1_identity(
        *,
        wb,
        header_font,
        header_fill,
        form_code: str,
        fiscal_year: int,
        profile,
    ) -> str:
        """Fiche R1 — taxpayer identity block (DGI standard)."""
        from openpyxl.styles import Alignment, Font

        sheet_name = "Fiche R1 - Identity"
        ws = wb.create_sheet(sheet_name)

        # Title row
        ws.cell(row=1, column=1, value="FICHE R1 — IDENTIFICATION DU CONTRIBUABLE")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(row=2, column=1, value=f"Exercice / Fiscal year: {fiscal_year}")
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

        # Header row for the field/value table
        ws.cell(row=4, column=1, value="Champ / Field")
        ws.cell(row=4, column=2, value="Valeur / Value")
        for col in (1, 2):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _v(value) -> str:
            if value is None:
                return ""
            if isinstance(value, bool):
                return "Yes" if value else "No"
            return str(value)

        rows: list[tuple[str, str]] = [
            ("DSF form / Formulaire DSF", form_code),
            ("Fiscal year / Exercice", str(fiscal_year)),
        ]
        if profile is not None:
            rows.extend(
                [
                    ("NIU / Numéro d'Identifiant Unique", _v(profile.niu)),
                    ("Tax center / Centre des impôts", _v(profile.tax_center_code)),
                    (
                        "Taxpayer segment / Segment du contribuable",
                        _v(profile.taxpayer_segment_code),
                    ),
                    ("Tax regime / Régime fiscal", _v(profile.tax_regime_code)),
                    ("VAT liable / Assujetti TVA", _v(profile.is_vat_liable)),
                    (
                        "VAT effective from / TVA effective au",
                        _v(profile.vat_effective_from),
                    ),
                    (
                        "CIT rate profile / Profil IS",
                        _v(profile.cit_rate_profile_code),
                    ),
                    (
                        "CIT installment profile / Profil acomptes IS",
                        _v(profile.cit_installment_profile_code),
                    ),
                    ("SME qualified / PME qualifiée", _v(profile.sme_qualified_flag)),
                    (
                        "DSF submission mode / Mode de dépôt DSF",
                        _v(profile.dsf_submission_mode_code),
                    ),
                    (
                        "OTP enabled / Guichet électronique",
                        _v(profile.otp_enabled_flag),
                    ),
                    (
                        "Default withholding applicable / Retenue à la source par défaut",
                        _v(profile.default_withholding_applicable_flag),
                    ),
                ]
            )
        else:
            rows.append(("Tax profile", "(not configured)"))

        for offset, (label, value) in enumerate(rows, start=5):
            ws.cell(row=offset, column=1, value=label)
            ws.cell(row=offset, column=2, value=value)

        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 40
        ws.freeze_panes = "A5"
        return sheet_name

    @staticmethod
    def _write_fiche_r2_revenue(
        *,
        wb,
        write_header,
        profile,
        obligations: list,
        returns: list,
        fiscal_year: int,
    ) -> str:
        """Fiche R2 — turnover, VAT, and activity codes block.

        Builds the annual VAT roll-up from monthly returns: total taxable
        base (sum of declared VAT base lines), total declared VAT, and
        total VAT paid for the year. Activity-code rows are placeholder
        slots for the DGI ``FICHE_R2`` codes that the user fills in
        before filing (the codes are listed in section 10 of the DSF
        2025 Guide).
        """
        sheet_name = "Fiche R2 - Revenue"
        ws = wb.create_sheet(sheet_name)
        write_header(
            ws,
            ["Code / Code", "Libellé / Label", "Montant / Amount", "Notes"],
        )

        # Annual VAT roll-up from existing tax returns
        vat_returns = [r for r in returns if r.tax_type_code == TAX_TYPE_VAT]
        total_vat_due = sum((r.total_due_amount for r in vat_returns), Decimal("0"))
        total_vat_paid = sum((r.total_paid_amount for r in vat_returns), Decimal("0"))

        # Sum across all VAT return lines whose box code looks like a
        # taxable-base box (best-effort heuristic; DGI uses codes such
        # as ``BASE_19_25``, ``BASE_EXPORT``, ``BASE_EXEMPT``).
        total_taxable_base = Decimal("0")
        for r in vat_returns:
            for line in r.lines:
                code = (line.box_code or "").upper()
                if code.startswith("BASE_") or code.startswith("CA_"):
                    total_taxable_base += line.amount

        rows = [
            (
                "R2-CA-TOTAL",
                "Chiffre d'affaires total / Total turnover",
                float(total_taxable_base),
                "Sum of monthly VAT base box values",
            ),
            (
                "R2-VAT-DUE",
                "TVA due totale / Total VAT due",
                float(total_vat_due),
                f"Sum of {len(vat_returns)} monthly VAT return totals",
            ),
            (
                "R2-VAT-PAID",
                "TVA payée totale / Total VAT paid",
                float(total_vat_paid),
                "Sum of payments recorded against VAT returns",
            ),
        ]

        # Add placeholder activity-code rows (DGI section 10 codes).
        rows.extend(
            [
                (
                    "R2-ACT-PRIMARY",
                    "Code activité principale / Primary activity code",
                    "",
                    "To be set by preparer (DSF Guide §10)",
                ),
                (
                    "R2-ACT-SECONDARY",
                    "Code activité secondaire / Secondary activity code",
                    "",
                    "Optional (DSF Guide §10)",
                ),
                (
                    "R2-EMPLOYEES",
                    "Effectif moyen / Average headcount",
                    "",
                    "Sourced from payroll module (annual average)",
                ),
            ]
        )

        for code, label, amount, note in rows:
            ws.append([code, label, amount, note])

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 60
        return sheet_name

    @staticmethod
    def _write_fiche_r3_balance_sheet(
        *,
        wb,
        write_header,
        fiscal_year: int,
        amounts: dict[str, dict[str, Decimal]] | None = None,
    ) -> str:
        """Fiche R3 — balance sheet structure (OHADA/SYSCOHADA layout).

        When ``amounts`` is supplied, the writer populates the Gross /
        Depreciation / Net columns from the OHADA balance sheet
        computation (keyed by reference code, e.g. ``"AI"``, ``"BG"``,
        ``"CP"``). When ``amounts`` is ``None`` (no balance-sheet
        service injected, or computation failed), the cells are left
        blank and a readiness warning is recorded by the caller.
        """
        sheet_name = "Fiche R3 - Balance Sheet"
        ws = wb.create_sheet(sheet_name)
        write_header(
            ws,
            [
                "Réf / Ref",
                "Poste / Line item",
                "Brut / Gross",
                "Amort./Prov. / Depr.",
                "Net / Net",
                "Net N-1 / Prior",
            ],
        )

        # SYSCOHADA balance sheet structure (assets side then liabilities)
        asset_rows = [
            ("AD", "Charges immobilisées / Capitalised expenses"),
            ("AE", "Frais d'établissement / Formation expenses"),
            ("AF", "Charges à répartir / Deferred charges"),
            ("AG", "Primes de remboursement / Bond redemption premium"),
            ("AH", "Immobilisations incorporelles / Intangible assets"),
            ("AI", "Immobilisations corporelles / Tangible assets"),
            ("AN", "Immobilisations financières / Financial assets"),
            ("AZ", "TOTAL ACTIF IMMOBILISÉ / Total non-current assets"),
            ("BA", "Stocks / Inventories"),
            ("BB", "Créances et emplois assimilés / Receivables"),
            ("BG", "Trésorerie-Actif / Cash & equivalents"),
            ("BZ", "TOTAL ACTIF CIRCULANT / Total current assets"),
            ("BT", "Écart de conversion-Actif / FX translation - Asset"),
            ("CZ", "TOTAL GÉNÉRAL ACTIF / Total assets"),
        ]
        liability_rows = [
            ("CA", "Capital / Share capital"),
            ("CB", "Apporteurs, capital non appelé / Uncalled capital", ""),
            ("CD", "Primes et réserves / Premiums and reserves"),
            ("CK", "Résultat net de l'exercice / Net result of the year"),
            ("CP", "TOTAL CAPITAUX PROPRES / Total equity"),
            ("DA", "Emprunts et dettes financières / Borrowings"),
            ("DP", "TOTAL DETTES FINANCIÈRES / Total financial liabilities"),
            ("DH", "Fournisseurs d'exploitation / Trade payables"),
            ("DI", "Dettes fiscales et sociales / Tax & social liabilities"),
            ("DZ", "TOTAL PASSIF CIRCULANT / Total current liabilities"),
            ("EZ", "TOTAL GÉNÉRAL PASSIF / Total liabilities & equity"),
        ]

        ws.append(["", "ACTIF / ASSETS", "", "", "", ""])
        for ref, label, *_ in asset_rows:
            cells = DSFExportService._fiche_r3_amount_cells(amounts, ref)
            ws.append([ref, label, *cells])
        ws.append(["", "", "", "", "", ""])
        ws.append(["", "PASSIF / EQUITY & LIABILITIES", "", "", "", ""])
        for ref, label, *_ in liability_rows:
            cells = DSFExportService._fiche_r3_amount_cells(amounts, ref)
            ws.append([ref, label, *cells])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 56
        for col in ("C", "D", "E", "F"):
            ws.column_dimensions[col].width = 16
        return sheet_name

    @staticmethod
    def _fiche_r3_amount_cells(
        amounts: dict[str, dict[str, Decimal]] | None,
        ref: str,
    ) -> tuple:
        """Build the (Gross, Depr, Net, PriorNet) cell tuple for R3.

        Prior-year net is left blank — only the current period is
        computable from posted accounting truth without an additional
        prior-year close.
        """
        if amounts is None:
            return ("", "", "", "")
        line = amounts.get(ref)
        if line is None:
            return ("", "", "", "")
        gross = line.get("gross", _ZERO) or _ZERO
        contra = line.get("contra", _ZERO) or _ZERO
        net = line.get("net", _ZERO) or _ZERO
        if gross == _ZERO and contra == _ZERO and net == _ZERO:
            return ("", "", "", "")
        return (
            float(gross),
            float(contra),
            float(net),
            "",
        )

    @staticmethod
    def _write_fiche_income_statement(
        *,
        wb,
        write_header,
        fiscal_year: int,
        simplified: bool,
        amounts: dict[str, Decimal] | None = None,
    ) -> str:
        """Income statement (Fiche R4) — full or simplified layout.

        When ``amounts`` is supplied (keyed by OHADA reference code,
        e.g. ``"TA"``, ``"RA"``, ``"RI"``), the *Year N* column is
        populated. *Year N-1* is always left blank — that requires
        comparative-period reporting which lives outside this slice.
        """
        sheet_name = (
            "Fiche - Simplified P&L"
            if simplified
            else "Fiche R4 - Income Statement"
        )
        ws = wb.create_sheet(sheet_name)
        write_header(
            ws,
            [
                "Réf / Ref",
                "Poste / Line item",
                "Exercice N / Year N",
                "Exercice N-1 / Year N-1",
            ],
        )

        full_rows = [
            ("TA", "Ventes de marchandises / Sale of goods"),
            ("TB", "Achats de marchandises / Purchases of goods"),
            ("RA", "MARGE COMMERCIALE / Gross trading margin"),
            ("TC", "Ventes de produits fabriqués / Sale of finished goods"),
            ("TD", "Travaux, services vendus / Services rendered"),
            ("TE", "Production stockée / Change in inventory"),
            ("TF", "Production immobilisée / Self-constructed assets"),
            ("TG", "Subventions d'exploitation / Operating subsidies"),
            ("TH", "Autres produits / Other operating income"),
            ("RB", "VALEUR AJOUTÉE / Value added"),
            ("RC", "EXCÉDENT BRUT D'EXPLOITATION / EBITDA"),
            ("RD", "RÉSULTAT D'EXPLOITATION / Operating result"),
            ("RE", "RÉSULTAT FINANCIER / Financial result"),
            ("RF", "RÉSULTAT DES ACTIVITÉS ORDINAIRES / Result from ordinary activities"),
            ("RG", "RÉSULTAT HORS ACTIVITÉS ORDINAIRES / Result from non-ordinary activities"),
            ("RH", "Impôts sur le résultat / Income tax"),
            ("RI", "RÉSULTAT NET / Net result"),
        ]
        simple_rows = [
            ("S1", "Recettes / Receipts"),
            ("S2", "Achats consommés / Purchases consumed"),
            ("S3", "Charges de personnel / Personnel expenses"),
            ("S4", "Autres charges / Other charges"),
            ("S5", "Résultat avant impôt / Pre-tax result"),
            ("S6", "Impôt sur le résultat / Income tax"),
            ("S7", "RÉSULTAT NET / Net result"),
        ]
        for ref, label in (simple_rows if simplified else full_rows):
            value: object = ""
            if amounts is not None:
                amount = amounts.get(ref)
                if amount is not None and amount != _ZERO:
                    value = float(amount)
            ws.append([ref, label, value, ""])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        return sheet_name

    @staticmethod
    def _write_fiche_cit_summary(
        *,
        wb,
        write_header,
        fiscal_year: int,
        obligations: list,
        returns: list,
        payments: list,
    ) -> str:
        """CIT (corporate income tax) annual reconciliation summary.

        Aggregates the year's CIT installments and balance from
        ``tax_obligations``/``tax_returns``/``tax_payments`` so the
        preparer can reconcile against the income statement before
        filing.
        """
        sheet_name = "Fiche - CIT Summary"
        ws = wb.create_sheet(sheet_name)
        write_header(
            ws,
            [
                "Période / Period",
                "Type / Tax type",
                "Échéance / Due",
                "Statut obligation / Obl. status",
                "Statut déclaration / Return status",
                "Montant dû / Due",
                "Montant payé / Paid",
            ],
        )

        cit_types = (TAX_TYPE_CIT_INSTALLMENT, TAX_TYPE_CIT_BALANCE)
        cit_obligations = [o for o in obligations if o.tax_type_code in cit_types]
        returns_by_obligation = {r.obligation_id: r for r in returns}

        total_due = Decimal("0")
        total_paid = Decimal("0")

        for o in sorted(cit_obligations, key=lambda x: x.period_start):
            r = returns_by_obligation.get(o.id)
            due = r.total_due_amount if r else Decimal("0")
            paid = r.total_paid_amount if r else Decimal("0")
            total_due += due
            total_paid += paid
            ws.append(
                [
                    f"{o.period_start} → {o.period_end}",
                    o.tax_type_code,
                    o.due_date,
                    o.status_code,
                    r.status_code if r else "",
                    float(due),
                    float(paid),
                ]
            )

        # Totals row
        ws.append(["", "", "", "", "TOTAL", float(total_due), float(total_paid)])
        # Net balance (positive = outstanding)
        ws.append(
            ["", "", "", "", "NET BALANCE", float(total_due - total_paid), ""]
        )

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 16
        return sheet_name

    @staticmethod
    def _write_fiche_liberatory_summary(
        *,
        wb,
        write_header,
        fiscal_year: int,
        profile,
        obligations: list,
        returns: list,
        payments: list,
    ) -> str:
        """Liberatory / IGS regime summary.

        Liberatory taxpayers (small/CDI) file a single annual flat-rate
        statement with their NIU, segment, and the IGS / patente
        amounts paid during the year.
        """
        sheet_name = "Fiche - Liberatory Summary"
        ws = wb.create_sheet(sheet_name)
        write_header(ws, ["Field / Champ", "Value / Valeur"])

        def _v(value) -> str:
            if value is None:
                return ""
            if isinstance(value, bool):
                return "Yes" if value else "No"
            return str(value)

        ws.append(["Fiscal year / Exercice", str(fiscal_year)])
        ws.append(
            [
                "NIU",
                _v(getattr(profile, "niu", None)) if profile else "",
            ]
        )
        ws.append(
            [
                "Tax center / Centre des impôts",
                _v(getattr(profile, "tax_center_code", None)) if profile else "",
            ]
        )
        ws.append(
            [
                "Taxpayer segment / Segment",
                _v(getattr(profile, "taxpayer_segment_code", None))
                if profile
                else "",
            ]
        )
        ws.append(
            [
                "Tax regime / Régime fiscal",
                _v(getattr(profile, "tax_regime_code", None)) if profile else "",
            ]
        )

        # Payments roll-up (IGS, patente, et al. recorded as tax_payments)
        total_paid = sum((p.amount for p in payments), Decimal("0"))
        ws.append(["Total taxes paid / Total impôts payés", float(total_paid)])
        ws.append(["Number of payments / Nombre de paiements", len(payments)])

        # Number of filed obligations
        filed_obligation_count = sum(
            1
            for o in obligations
            if o.status_code in ("FILED", "PAID")
        )
        ws.append(
            [
                "Filed obligations / Obligations déposées",
                filed_obligation_count,
            ]
        )

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 32
        return sheet_name

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _validate_year(year: int) -> None:
        if not isinstance(year, int):
            raise ValidationError("Fiscal year must be an integer.")
        if year < 2000 or year > 2100:
            raise ValidationError("Fiscal year is outside the supported range.")

    def _require_company_exists(self, session: Session, company_id: int) -> None:
        company_repo = self._company_repository_factory(session)
        if company_repo.get_by_id(company_id) is None:
            raise NotFoundError(f"Company with id {company_id} was not found.")

    def _record_audit(self, company_id: int, description: str) -> None:
        if self._audit_service is None:
            return
        from seeker_accounting.modules.audit.dto.audit_event_dto import (
            RecordAuditEventCommand,
        )
        from seeker_accounting.modules.audit.event_type_catalog import MODULE_TAXATION

        try:
            self._audit_service.record_event(
                company_id,
                RecordAuditEventCommand(
                    event_type_code="DSF_EXPORT_GENERATED",
                    module_code=MODULE_TAXATION,
                    entity_type="DSFExport",
                    entity_id=None,
                    description=description,
                ),
            )
        except Exception:
            pass
