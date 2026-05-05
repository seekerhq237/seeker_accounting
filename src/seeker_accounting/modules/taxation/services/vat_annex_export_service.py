"""VAT annex export service (Slice T48).

Produces the two DGI annexes required when annual customer or supplier
VAT-bearing turnover exceeds the DGI disclosure threshold:

* **Customers annex** (`états annexes clients`) — per-customer taxable
  base and output VAT for a given calendar/fiscal month, keyed on the
  customer NIU (``tax_identifier``).
* **Suppliers annex** (`états annexes fournisseurs`) — per-supplier
  taxable base and deductible input VAT for the same period.

Design principles
-----------------
* Reads from ``posted_tax_lines`` joined to the source documents and
  the customer/supplier master tables via raw SQL (SQLAlchemy Core
  ``select``).  Keeps join logic isolated here; no new repository
  methods are required.
* Produces ``bytes`` (an openpyxl workbook serialised to a BytesIO
  buffer).  The caller decides where to save the file.
* Respects company scoping.  No cross-company data leakage possible.
* Permission-gated: requires ``taxation.returns.export_annex``.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING, Callable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from seeker_accounting.app.context.app_context import AppContext
from seeker_accounting.db.unit_of_work import UnitOfWorkFactory
from seeker_accounting.modules.administration.services.permission_service import (
    PermissionService,
)
from seeker_accounting.modules.companies.repositories.company_repository import (
    CompanyRepository,
)
from seeker_accounting.modules.customers.models.customer import Customer
from seeker_accounting.modules.purchases.models.purchase_bill import PurchaseBill
from seeker_accounting.modules.purchases.models.purchase_credit_note import (
    PurchaseCreditNote,
)
from seeker_accounting.modules.sales.models.sales_credit_note import SalesCreditNote
from seeker_accounting.modules.sales.models.sales_invoice import SalesInvoice
from seeker_accounting.modules.suppliers.models.supplier import Supplier
from seeker_accounting.modules.taxation.models.posted_tax_line import (
    DIRECTION_PURCHASE,
    DIRECTION_SALES,
    SOURCE_PURCHASE_BILL,
    SOURCE_PURCHASE_CREDIT_NOTE,
    SOURCE_SALES_CREDIT_NOTE,
    SOURCE_SALES_INVOICE,
    PostedTaxLine,
)
from seeker_accounting.platform.exceptions import NotFoundError

if TYPE_CHECKING:
    pass


PERMISSION_EXPORT_ANNEX: str = "taxation.returns.export_annex"

# Default OHADA DGI disclosure threshold (XAF annual turnover).
# Companies below this can still generate annexes; the field is
# informational and not enforced here.
DEFAULT_ANNEX_THRESHOLD_XAF: Decimal = Decimal("50000000")

# ── Column headers ────────────────────────────────────────────────────────────

_CUSTOMERS_HEADERS: list[str] = [
    "NIU Client",
    "Nom / Raison sociale",
    "Nombre de factures",
    "Base HT totale (FCFA)",
    "TVA collectée (FCFA)",
    "Dont TVA retenue à la source (FCFA)",
]

_SUPPLIERS_HEADERS: list[str] = [
    "NIU Fournisseur",
    "Nom / Raison sociale",
    "Nombre de factures",
    "Base HT totale (FCFA)",
    "TVA déductible (FCFA)",
]


# ── DTOs ─────────────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class AnnexLineDTO:
    """One aggregated counterparty row in an annex."""

    tax_identifier: str | None
    display_name: str
    document_count: int
    taxable_base: Decimal
    tax_amount: Decimal
    withholding_vat_amount: Decimal = Decimal("0")


@dataclass(frozen=True)
class VATAnnexResultDTO:
    """Result bundle returned by the export service."""

    company_name: str
    period_start: date
    period_end: date
    direction: str  # SALES | PURCHASE
    rows: list[AnnexLineDTO]
    workbook_bytes: bytes


# ── Internal aggregate helpers ────────────────────────────────────────────────


def _aggregate_sales(
    session: Session,
    company_id: int,
    period_start: date,
    period_end: date,
) -> list[AnnexLineDTO]:
    """Aggregate output VAT per customer for the date range.

    Joins ``posted_tax_lines`` → ``sales_invoices`` or
    ``sales_credit_notes`` → ``customers``.
    """
    # Sales invoices aggregation
    inv_stmt = (
        select(
            Customer.id.label("customer_id"),
            Customer.display_name.label("display_name"),
            Customer.tax_identifier.label("tax_identifier"),
            func.count(func.distinct(SalesInvoice.id)).label("doc_count"),
            func.coalesce(func.sum(PostedTaxLine.taxable_base), 0).label("base"),
            func.coalesce(func.sum(PostedTaxLine.tax_amount), 0).label("tax"),
        )
        .join(SalesInvoice, PostedTaxLine.source_document_id == SalesInvoice.id)
        .join(Customer, SalesInvoice.customer_id == Customer.id)
        .where(
            PostedTaxLine.company_id == company_id,
            PostedTaxLine.direction == DIRECTION_SALES,
            PostedTaxLine.source_document_type == SOURCE_SALES_INVOICE,
            PostedTaxLine.tax_point_date >= period_start,
            PostedTaxLine.tax_point_date <= period_end,
        )
        .group_by(Customer.id, Customer.display_name, Customer.tax_identifier)
    )

    # Credit notes aggregation (amounts are negative in posted_tax_lines)
    cn_stmt = (
        select(
            Customer.id.label("customer_id"),
            Customer.display_name.label("display_name"),
            Customer.tax_identifier.label("tax_identifier"),
            func.count(func.distinct(SalesCreditNote.id)).label("doc_count"),
            func.coalesce(func.sum(PostedTaxLine.taxable_base), 0).label("base"),
            func.coalesce(func.sum(PostedTaxLine.tax_amount), 0).label("tax"),
        )
        .join(
            SalesCreditNote,
            PostedTaxLine.source_document_id == SalesCreditNote.id,
        )
        .join(Customer, SalesCreditNote.customer_id == Customer.id)
        .where(
            PostedTaxLine.company_id == company_id,
            PostedTaxLine.direction == DIRECTION_SALES,
            PostedTaxLine.source_document_type == SOURCE_SALES_CREDIT_NOTE,
            PostedTaxLine.tax_point_date >= period_start,
            PostedTaxLine.tax_point_date <= period_end,
        )
        .group_by(Customer.id, Customer.display_name, Customer.tax_identifier)
    )

    # Merge the two result sets
    totals: dict[int, dict] = {}

    for row in session.execute(inv_stmt):
        totals[row.customer_id] = {
            "display_name": row.display_name,
            "tax_identifier": row.tax_identifier,
            "doc_count": row.doc_count,
            "base": Decimal(str(row.base)),
            "tax": Decimal(str(row.tax)),
        }

    for row in session.execute(cn_stmt):
        cid = row.customer_id
        if cid in totals:
            totals[cid]["doc_count"] += row.doc_count
            totals[cid]["base"] += Decimal(str(row.base))
            totals[cid]["tax"] += Decimal(str(row.tax))
        else:
            totals[cid] = {
                "display_name": row.display_name,
                "tax_identifier": row.tax_identifier,
                "doc_count": row.doc_count,
                "base": Decimal(str(row.base)),
                "tax": Decimal(str(row.tax)),
            }

    return [
        AnnexLineDTO(
            tax_identifier=v["tax_identifier"],
            display_name=v["display_name"],
            document_count=v["doc_count"],
            taxable_base=v["base"],
            tax_amount=v["tax"],
        )
        for v in sorted(totals.values(), key=lambda x: x["display_name"])
    ]


def _aggregate_purchases(
    session: Session,
    company_id: int,
    period_start: date,
    period_end: date,
) -> list[AnnexLineDTO]:
    """Aggregate deductible input VAT per supplier for the date range."""

    bill_stmt = (
        select(
            Supplier.id.label("supplier_id"),
            Supplier.display_name.label("display_name"),
            Supplier.tax_identifier.label("tax_identifier"),
            func.count(func.distinct(PurchaseBill.id)).label("doc_count"),
            func.coalesce(func.sum(PostedTaxLine.taxable_base), 0).label("base"),
            func.coalesce(
                func.sum(
                    PostedTaxLine.tax_amount
                ),
                0,
            ).label("tax"),
        )
        .join(PurchaseBill, PostedTaxLine.source_document_id == PurchaseBill.id)
        .join(Supplier, PurchaseBill.supplier_id == Supplier.id)
        .where(
            PostedTaxLine.company_id == company_id,
            PostedTaxLine.direction == DIRECTION_PURCHASE,
            PostedTaxLine.source_document_type == SOURCE_PURCHASE_BILL,
            PostedTaxLine.is_recoverable == True,  # noqa: E712 — SQLAlchemy comparison
            PostedTaxLine.tax_point_date >= period_start,
            PostedTaxLine.tax_point_date <= period_end,
        )
        .group_by(Supplier.id, Supplier.display_name, Supplier.tax_identifier)
    )

    cn_stmt = (
        select(
            Supplier.id.label("supplier_id"),
            Supplier.display_name.label("display_name"),
            Supplier.tax_identifier.label("tax_identifier"),
            func.count(func.distinct(PurchaseCreditNote.id)).label("doc_count"),
            func.coalesce(func.sum(PostedTaxLine.taxable_base), 0).label("base"),
            func.coalesce(func.sum(PostedTaxLine.tax_amount), 0).label("tax"),
        )
        .join(
            PurchaseCreditNote,
            PostedTaxLine.source_document_id == PurchaseCreditNote.id,
        )
        .join(Supplier, PurchaseCreditNote.supplier_id == Supplier.id)
        .where(
            PostedTaxLine.company_id == company_id,
            PostedTaxLine.direction == DIRECTION_PURCHASE,
            PostedTaxLine.source_document_type == SOURCE_PURCHASE_CREDIT_NOTE,
            PostedTaxLine.is_recoverable == True,  # noqa: E712
            PostedTaxLine.tax_point_date >= period_start,
            PostedTaxLine.tax_point_date <= period_end,
        )
        .group_by(Supplier.id, Supplier.display_name, Supplier.tax_identifier)
    )

    totals: dict[int, dict] = {}

    for row in session.execute(bill_stmt):
        totals[row.supplier_id] = {
            "display_name": row.display_name,
            "tax_identifier": row.tax_identifier,
            "doc_count": row.doc_count,
            "base": Decimal(str(row.base)),
            "tax": Decimal(str(row.tax)),
        }

    for row in session.execute(cn_stmt):
        sid = row.supplier_id
        if sid in totals:
            totals[sid]["doc_count"] += row.doc_count
            totals[sid]["base"] += Decimal(str(row.base))
            totals[sid]["tax"] += Decimal(str(row.tax))
        else:
            totals[sid] = {
                "display_name": row.display_name,
                "tax_identifier": row.tax_identifier,
                "doc_count": row.doc_count,
                "base": Decimal(str(row.base)),
                "tax": Decimal(str(row.tax)),
            }

    return [
        AnnexLineDTO(
            tax_identifier=v["tax_identifier"],
            display_name=v["display_name"],
            document_count=v["doc_count"],
            taxable_base=v["base"],
            tax_amount=v["tax"],
        )
        for v in sorted(totals.values(), key=lambda x: x["display_name"])
    ]


# ── Workbook renderer ─────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F497D")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FONT = Font(bold=True)


def _build_workbook(
    title: str,
    company_name: str,
    period_start: date,
    period_end: date,
    headers: list[str],
    rows: list[AnnexLineDTO],
    *,
    is_sales: bool,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit

    # Title row
    ws.append([f"ÉTAT ANNEXE — {title.upper()} — {company_name}"])
    ws.append([f"Période : {period_start.strftime('%d/%m/%Y')} – {period_end.strftime('%d/%m/%Y')}"])
    ws.append([])

    # Header row
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    total_base = Decimal("0")
    total_tax = Decimal("0")
    total_docs = 0

    for line in rows:
        if is_sales:
            ws.append([
                line.tax_identifier or "",
                line.display_name,
                line.document_count,
                float(line.taxable_base),
                float(line.tax_amount),
                float(line.withholding_vat_amount),
            ])
        else:
            ws.append([
                line.tax_identifier or "",
                line.display_name,
                line.document_count,
                float(line.taxable_base),
                float(line.tax_amount),
            ])
        total_base += line.taxable_base
        total_tax += line.tax_amount
        total_docs += line.document_count

    # Totals row
    if is_sales:
        total_row = ["", "TOTAL", total_docs, float(total_base), float(total_tax), ""]
    else:
        total_row = ["", "TOTAL", total_docs, float(total_base), float(total_tax)]
    ws.append(total_row)
    total_data_row = ws.max_row
    for cell in ws[total_data_row]:
        cell.font = _TOTAL_FONT

    # Column widths
    col_widths = [20, 40, 12, 22, 22, 22]
    for i, w in enumerate(col_widths[: ws.max_column], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Right-align numeric columns (3 onward)
    num_align = Alignment(horizontal="right")
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row[2:]:
            cell.alignment = num_align

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Service ───────────────────────────────────────────────────────────────────


class VATAnnexExportService:
    """Generates VAT annex Excel files for a given period.

    Both ``generate_customers_annex`` and ``generate_suppliers_annex``
    accept a half-open date range ``[period_start, period_end]`` and
    return a ``VATAnnexResultDTO`` that includes the raw workbook bytes.
    """

    PERMISSION_EXPORT_ANNEX: str = PERMISSION_EXPORT_ANNEX

    def __init__(
        self,
        unit_of_work_factory: UnitOfWorkFactory,
        company_repository_factory: Callable[[Session], CompanyRepository],
        permission_service: PermissionService,
        app_context: AppContext | None = None,
    ) -> None:
        self._uow_factory = unit_of_work_factory
        self._company_repo_factory = company_repository_factory
        self._permission_service = permission_service
        self._app_context = app_context

    # ── Public methods ────────────────────────────────────────────────────────

    def generate_customers_annex(
        self,
        company_id: int,
        period_start: date,
        period_end: date,
    ) -> VATAnnexResultDTO:
        """Build the customers (output VAT) annex workbook."""
        self._permission_service.require_permission(self.PERMISSION_EXPORT_ANNEX)
        with self._uow_factory() as session:
            company = self._company_repo_factory(session).get(company_id)
            if company is None:
                raise NotFoundError(f"Company {company_id} not found.")
            rows = _aggregate_sales(session, company_id, period_start, period_end)
            wb_bytes = _build_workbook(
                "Clients",
                company.display_name,
                period_start,
                period_end,
                _CUSTOMERS_HEADERS,
                rows,
                is_sales=True,
            )
        return VATAnnexResultDTO(
            company_name=company.display_name,
            period_start=period_start,
            period_end=period_end,
            direction=DIRECTION_SALES,
            rows=rows,
            workbook_bytes=wb_bytes,
        )

    def generate_suppliers_annex(
        self,
        company_id: int,
        period_start: date,
        period_end: date,
    ) -> VATAnnexResultDTO:
        """Build the suppliers (input VAT) annex workbook."""
        self._permission_service.require_permission(self.PERMISSION_EXPORT_ANNEX)
        with self._uow_factory() as session:
            company = self._company_repo_factory(session).get(company_id)
            if company is None:
                raise NotFoundError(f"Company {company_id} not found.")
            rows = _aggregate_purchases(
                session, company_id, period_start, period_end
            )
            wb_bytes = _build_workbook(
                "Fournisseurs",
                company.display_name,
                period_start,
                period_end,
                _SUPPLIERS_HEADERS,
                rows,
                is_sales=False,
            )
        return VATAnnexResultDTO(
            company_name=company.display_name,
            period_start=period_start,
            period_end=period_end,
            direction=DIRECTION_PURCHASE,
            rows=rows,
            workbook_bytes=wb_bytes,
        )
