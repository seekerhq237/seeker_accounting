"""Smoke script for the Tax Compliance workspace (Slice T7)."""

from __future__ import annotations

import os
import sys
import tempfile
import time
from datetime import date
from decimal import Decimal

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication

from shared.bootstrap import bootstrap_script_runtime
from seeker_accounting.app.navigation import nav_ids
from seeker_accounting.app.shell.main_window import MainWindow
from seeker_accounting.modules.accounting.reference_data.models.country import Country
from seeker_accounting.modules.accounting.reference_data.models.currency import Currency
from seeker_accounting.modules.companies.dto.company_commands import (
    CreateCompanyCommand,
)
from seeker_accounting.modules.taxation.dto.dsf_export_dto import (
    GenerateDSFExportCommand,
)
from seeker_accounting.modules.taxation.dto.tax_compliance_dto import (
    GenerateMonthlyVATObligationsCommand,
    GenerateQuarterlyCITInstallmentsCommand,
)
from seeker_accounting.modules.taxation.dto.company_tax_profile_dto import (
    UpsertCompanyTaxProfileCommand,
)
from seeker_accounting.modules.taxation.ui.tax_compliance_dialogs import (
    DSFExportDialog,
    GenerateMonthlyVATObligationsDialog,
    GenerateQuarterlyCITInstallmentsDialog,
)
from seeker_accounting.modules.taxation.ui.tax_compliance_page import (
    TaxCompliancePage,
)


def _get_page(main_window, nav_id, page_class, app):
    main_window._service_registry.navigation_service.navigate(nav_id)
    app.processEvents()
    page = main_window.findChild(page_class)
    if page is None:
        raise SystemExit(
            f"Expected {page_class.__name__} for nav_id {nav_id} not found"
        )
    return page


def main() -> int:
    app = QApplication([])
    from seeker_accounting.modules.administration.rbac_catalog import (
        ALL_SYSTEM_PERMISSION_CODES,
    )

    bootstrap = bootstrap_script_runtime(
        app, permission_snapshot=ALL_SYSTEM_PERMISSION_CODES
    )
    registry = bootstrap.service_registry

    with registry.session_context.unit_of_work_factory() as uow:
        session = uow.session
        if session.get(Country, "CM") is None:
            session.add(Country(code="CM", name="Cameroon", is_active=True))
        if session.get(Currency, "XAF") is None:
            session.add(
                Currency(
                    code="XAF",
                    name="Central African CFA franc",
                    symbol="FCFA",
                    decimal_places=0,
                    is_active=True,
                )
            )
        uow.commit()

    company = registry.company_service.create_company(
        CreateCompanyCommand(
            legal_name=f"Smoke Tax Compliance Co {int(time.time())}",
            display_name="Smoke Tax Compliance Co",
            country_code="CM",
            base_currency_code="XAF",
        )
    )
    registry.company_context_service.set_active_company(company.id)

    main_window = MainWindow(registry)
    main_window.show()
    app.processEvents()

    page = _get_page(main_window, nav_ids.TAX_COMPLIANCE, TaxCompliancePage, app)
    print(
        "tax_compliance_loaded_with_active",
        page._stack.currentWidget() is page._workspace,
    )
    print("initial_obligation_count", len(page._obligations))
    print("initial_return_count", len(page._returns))

    # Drive the obligation generator service directly (dialog wraps this).
    year = date.today().year
    generated = registry.tax_obligation_service.generate_monthly_vat_obligations(
        company.id,
        GenerateMonthlyVATObligationsCommand(year=year, due_day_of_next_month=15),
    )
    print("generated_count", len(generated))
    page.reload()
    app.processEvents()
    print("after_generate_obligation_count", len(page._obligations))
    print(
        "obligations_table_row_count",
        page._obligations_table.rowCount(),
    )

    # Briefly open and close each dialog to confirm they construct cleanly.
    gen_dialog = GenerateMonthlyVATObligationsDialog(registry, company.id, parent=page)
    gen_dialog.show()
    app.processEvents()
    gen_dialog.reject()
    app.processEvents()

    # Drive the CIT installment generator end-to-end.
    cit_generated = (
        registry.tax_obligation_service.generate_quarterly_cit_installments(
            company.id,
            GenerateQuarterlyCITInstallmentsCommand(
                year=year, due_day_of_next_month=15
            ),
        )
    )
    print("cit_generated_count", len(cit_generated))
    print("cit_first_period_start", cit_generated[0].period_start.isoformat())
    print("cit_last_due_date", cit_generated[-1].due_date.isoformat())
    page.reload()
    app.processEvents()
    print("after_cit_obligation_count", len(page._obligations))

    # Briefly open the CIT dialog as well.
    cit_dialog = GenerateQuarterlyCITInstallmentsDialog(
        registry, company.id, parent=page
    )
    cit_dialog.show()
    app.processEvents()
    cit_dialog.reject()
    app.processEvents()

    # Verify ribbon plumbing: surfaces are registered and the page exposes commands.
    ribbon_registry = registry.ribbon_registry
    print("ribbon_has_tax_profile", ribbon_registry.has("tax_profile"))
    print("ribbon_has_tax_compliance", ribbon_registry.has("tax_compliance"))
    print(
        "ribbon_command_count",
        len(page._ribbon_commands()),
    )

    dsf_dialog = DSFExportDialog(
        registry, company.id, company.display_name, parent=page
    )
    dsf_dialog.show()
    app.processEvents()
    # Run a readiness check through the dialog's button handler.
    dsf_dialog._handle_check_readiness()
    app.processEvents()
    print("dsf_readiness_summary", dsf_dialog._readiness_summary.text())
    dsf_dialog.reject()
    app.processEvents()

    # Drive the DSF export service end-to-end.
    with tempfile.TemporaryDirectory() as tmp:
        out_path = os.path.join(tmp, f"smoke_dsf_{year}.xlsx")
        result = registry.dsf_export_service.generate(
            company.id,
            GenerateDSFExportCommand(fiscal_year=year, output_path=out_path),
        )
        print("dsf_sheets_written", ",".join(result.sheets_written))
        print("dsf_obligation_count", result.obligation_count)
        print("dsf_file_exists", os.path.exists(out_path))
        print("dsf_form_applied", result.dsf_form_applied)
        print("dsf_regime_applied", result.tax_regime_applied)

        # Phase 4 — verify regime-conditional fiches landed in the
        # workbook for whichever DSF form the active company has.
        import openpyxl

        wb = openpyxl.load_workbook(out_path)
        sheet_names = set(wb.sheetnames)
        print("dsf_sheet_count", len(sheet_names))
        if result.dsf_form_applied == "DSF_REAL":
            for expected in (
                "Fiche R1 - Identity",
                "Fiche R2 - Revenue",
                "Fiche R3 - Balance Sheet",
                "Fiche R4 - Income Statement",
                "Fiche - CIT Summary",
            ):
                assert expected in sheet_names, f"missing fiche: {expected}"
            print("dsf_real_fiches_ok", True)
        elif result.dsf_form_applied == "DSF_SIMPLIFIED":
            for expected in (
                "Fiche R1 - Identity",
                "Fiche R2 - Revenue",
                "Fiche - Simplified P&L",
                "Fiche - CIT Summary",
            ):
                assert expected in sheet_names, f"missing fiche: {expected}"
            print("dsf_simplified_fiches_ok", True)
        elif result.dsf_form_applied == "DSF_LIBERATORY":
            for expected in (
                "Fiche R1 - Identity",
                "Fiche - Liberatory Summary",
            ):
                assert expected in sheet_names, f"missing fiche: {expected}"
            print("dsf_liberatory_fiches_ok", True)
        else:
            print("dsf_no_form_family_applied", True)

    # Phase 4 — drive the export across all three regime form families
    # to prove the regime-conditional fiche layouts work end-to-end.
    expected_fiches_by_form = {
        "DSF_REAL": (
            "Fiche R1 - Identity",
            "Fiche R2 - Revenue",
            "Fiche R3 - Balance Sheet",
            "Fiche R4 - Income Statement",
            "Fiche - CIT Summary",
        ),
        "DSF_SIMPLIFIED": (
            "Fiche R1 - Identity",
            "Fiche R2 - Revenue",
            "Fiche - Simplified P&L",
            "Fiche - CIT Summary",
        ),
        "DSF_LIBERATORY": (
            "Fiche R1 - Identity",
            "Fiche - Liberatory Summary",
        ),
    }
    regime_by_form = {
        "DSF_REAL": "REAL",
        "DSF_SIMPLIFIED": "SIMPLIFIED",
        "DSF_LIBERATORY": "LIBERATORY",
    }
    for form_code, expected_sheets in expected_fiches_by_form.items():
        registry.company_tax_profile_service.upsert(
            company.id,
            UpsertCompanyTaxProfileCommand(
                niu="P012345678901A",
                tax_center_code="DPMI_DOUALA",
                taxpayer_segment_code="MEDIUM",
                tax_regime_code=regime_by_form[form_code],
                is_vat_liable=True,
                vat_effective_from=date(year, 1, 1),
                cit_rate_profile_code="STANDARD",
                cit_installment_profile_code="QUARTERLY",
                sme_qualified_flag=False,
                dsf_form_code=form_code,
                dsf_submission_mode_code="EXCEL",
                otp_enabled_flag=False,
                default_withholding_applicable_flag=False,
            ),
        )
        with tempfile.TemporaryDirectory() as tmp:
            out_path = os.path.join(tmp, f"smoke_dsf_{form_code}.xlsx")
            r = registry.dsf_export_service.generate(
                company.id,
                GenerateDSFExportCommand(fiscal_year=year, output_path=out_path),
            )
            assert r.dsf_form_applied == form_code, (
                f"expected {form_code}, got {r.dsf_form_applied}"
            )
            assert r.tax_regime_applied == regime_by_form[form_code]
            import openpyxl

            wb = openpyxl.load_workbook(out_path)
            sheets = set(wb.sheetnames)
            for expected in expected_sheets:
                assert expected in sheets, (
                    f"{form_code}: missing fiche {expected!r}; got {sorted(sheets)}"
                )
            print(f"dsf_phase4_{form_code.lower()}_fiches_ok", True)

    # Selection reactivity on returns table (no returns yet => disabled file/payment).
    print("file_button_enabled_no_selection", page._file_button.isEnabled())
    print("payment_button_enabled_no_selection", page._payment_button.isEnabled())

    # Clear active company → page collapses to no-company card.
    registry.company_context_service.clear_active_company()
    app.processEvents()
    page.reload()
    app.processEvents()
    print(
        "tax_compliance_no_active",
        page._stack.currentWidget() is page._no_company_card,
    )

    main_window.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
