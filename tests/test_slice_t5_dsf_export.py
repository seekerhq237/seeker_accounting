"""Slice T5 tests — DSF export service.

Verifies the export workbook structure, readiness validation, and
permission gating. Uses mocked repositories (consistent with T4
service-test pattern).
"""

from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock

from seeker_accounting.db import model_registry  # noqa: F401

from seeker_accounting.modules.taxation.constants import (
    OBLIGATION_STATUS_OPEN,
    OBLIGATION_STATUS_PAID,
    RETURN_STATUS_DRAFT,
    RETURN_STATUS_FILED,
    TAX_PAYMENT_METHOD_BANK_TRANSFER,
    TAX_TYPE_VAT,
    VAT_BOX_INPUT_TAX_DEDUCTIBLE,
    VAT_BOX_NET_VAT_DUE,
    VAT_BOX_OUTPUT_TAX,
)
from seeker_accounting.modules.taxation.dto.dsf_export_dto import (
    GenerateDSFExportCommand,
)
from seeker_accounting.modules.taxation.services.dsf_export_service import (
    DSFExportService,
)
from seeker_accounting.platform.exceptions import (
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)


class _FakeUnitOfWork:
    def __init__(self) -> None:
        self.session = MagicMock(name="Session")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return None

    def commit(self):
        pass


class _FakePermissionService:
    def __init__(self, granted: set[str]) -> None:
        self._granted = granted

    def require_permission(self, code: str) -> None:
        if code not in self._granted:
            raise PermissionDeniedError(f"Missing permission: {code}")


def _build_profile(**overrides) -> SimpleNamespace:
    base = dict(
        company_id=1,
        niu="P012345678901A",
        tax_center_code="DPMI_DOUALA",
        taxpayer_segment_code="MEDIUM",
        tax_regime_code="REAL",
        is_vat_liable=True,
        vat_effective_from=date(2024, 1, 1),
        cit_rate_profile_code="STANDARD_33",
        cit_installment_profile_code="QUARTERLY",
        sme_qualified_flag=False,
        dsf_form_code="DSF_NORMAL",
        dsf_submission_mode_code="DGI_PORTAL",
        otp_enabled_flag=True,
        default_withholding_applicable_flag=False,
    )
    base.update(overrides)
    return SimpleNamespace(**base)


def _build_obligation(period_start: date, period_end: date, **overrides):
    base = dict(
        id=overrides.get("id", hash((period_start, period_end)) & 0xFFFF),
        company_id=1,
        tax_type_code=TAX_TYPE_VAT,
        period_start=period_start,
        period_end=period_end,
        due_date=overrides.get("due_date", period_end),
        status_code=OBLIGATION_STATUS_OPEN,
    )
    base.update(overrides)
    return SimpleNamespace(**base)


def _build_return(period_start: date, period_end: date, lines=(), **overrides):
    base = dict(
        id=overrides.get("id", 1),
        company_id=1,
        obligation_id=overrides.get("obligation_id", 100),
        tax_type_code=TAX_TYPE_VAT,
        period_start=period_start,
        period_end=period_end,
        status_code=RETURN_STATUS_DRAFT,
        total_due_amount=Decimal("0.00"),
        total_paid_amount=Decimal("0.00"),
        filed_at=None,
        otp_reference=None,
        external_reference=None,
        lines=list(lines),
        payments=[],
    )
    base.update(overrides)
    return SimpleNamespace(**base)


def _build_line(box_code: str, amount: Decimal, label: str = "", sort_order: int = 0):
    return SimpleNamespace(
        box_code=box_code,
        label=label or box_code,
        amount=amount,
        sort_order=sort_order,
    )


def _build_payment(payment_date: date, amount: Decimal, **overrides):
    base = dict(
        id=overrides.get("id", 1),
        company_id=1,
        tax_return_id=overrides.get("tax_return_id", 1),
        payment_date=payment_date,
        amount=amount,
        payment_method_code=TAX_PAYMENT_METHOD_BANK_TRANSFER,
        reference="REF-001",
        notes=None,
    )
    base.update(overrides)
    return SimpleNamespace(**base)


def _build_service(
    *,
    granted: set[str] | None = None,
    company_exists: bool = True,
    profile=None,
    obligations: list | None = None,
    returns: list | None = None,
    payments: list | None = None,
    ohada_balance_sheet_service=None,
    ohada_income_statement_service=None,
):
    if granted is None:
        granted = {DSFExportService.PERMISSION_EXPORT}

    uow = _FakeUnitOfWork()

    profile_repo = MagicMock()
    profile_repo.get_by_company.return_value = profile

    obligation_repo = MagicMock()
    obligation_repo.list_by_company.return_value = obligations or []

    return_repo = MagicMock()
    return_repo.list_by_company.return_value = returns or []
    by_id = {r.id: r for r in (returns or [])}
    return_repo.get_by_id.side_effect = lambda company_id, return_id: by_id.get(
        return_id
    )

    payment_repo = MagicMock()
    payment_repo.list_by_company.return_value = payments or []

    company_repo = MagicMock()
    company_repo.get_by_id.return_value = (
        SimpleNamespace(id=1, name="Acme") if company_exists else None
    )

    service = DSFExportService(
        unit_of_work_factory=lambda: uow,
        app_context=SimpleNamespace(current_user_id=42),
        company_tax_profile_repository_factory=lambda s: profile_repo,
        tax_obligation_repository_factory=lambda s: obligation_repo,
        tax_return_repository_factory=lambda s: return_repo,
        tax_payment_repository_factory=lambda s: payment_repo,
        company_repository_factory=lambda s: company_repo,
        permission_service=_FakePermissionService(granted),
        audit_service=None,
        ohada_balance_sheet_service=ohada_balance_sheet_service,
        ohada_income_statement_service=ohada_income_statement_service,
    )
    return service


class DSFExportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.output_path = os.path.join(self._tmpdir.name, "dsf_2025.xlsx")

    # ── Permission ────────────────────────────────────────────────

    def test_generate_requires_export_permission(self) -> None:
        service = _build_service(granted=set())
        with self.assertRaises(PermissionDeniedError):
            service.generate(
                1, GenerateDSFExportCommand(2025, self.output_path)
            )

    def test_check_readiness_requires_export_permission(self) -> None:
        service = _build_service(granted=set())
        with self.assertRaises(PermissionDeniedError):
            service.check_readiness(1, 2025)

    # ── Validation ────────────────────────────────────────────────

    def test_invalid_year_rejected(self) -> None:
        service = _build_service()
        with self.assertRaises(ValidationError):
            service.generate(1, GenerateDSFExportCommand(1999, self.output_path))

    def test_output_path_must_be_xlsx(self) -> None:
        service = _build_service()
        with self.assertRaises(ValidationError):
            service.generate(
                1, GenerateDSFExportCommand(2025, "/tmp/dsf.csv")
            )

    def test_output_path_required(self) -> None:
        service = _build_service()
        with self.assertRaises(ValidationError):
            service.generate(1, GenerateDSFExportCommand(2025, "   "))

    def test_unknown_company_raises_not_found(self) -> None:
        service = _build_service(company_exists=False)
        with self.assertRaises(NotFoundError):
            service.generate(
                1, GenerateDSFExportCommand(2025, self.output_path)
            )

    # ── Readiness ─────────────────────────────────────────────────

    def test_readiness_flags_missing_profile(self) -> None:
        service = _build_service(profile=None)
        issues = service.check_readiness(1, 2025)
        codes = {i.code for i in issues}
        self.assertIn("NO_TAX_PROFILE", codes)

    def test_readiness_flags_missing_niu(self) -> None:
        service = _build_service(profile=_build_profile(niu=""))
        issues = service.check_readiness(1, 2025)
        codes = {i.code for i in issues}
        self.assertIn("MISSING_NIU", codes)

    def test_readiness_flags_incomplete_vat_obligations(self) -> None:
        # VAT-liable but only 6 obligations
        obligations = [
            _build_obligation(date(2025, m, 1), date(2025, m, 28), id=m)
            for m in range(1, 7)
        ]
        service = _build_service(
            profile=_build_profile(),
            obligations=obligations,
        )
        issues = service.check_readiness(1, 2025)
        codes = {i.code for i in issues}
        self.assertIn("INCOMPLETE_VAT_OBLIGATIONS", codes)

    def test_readiness_flags_unfiled_returns(self) -> None:
        ret = _build_return(date(2025, 1, 1), date(2025, 1, 31))
        service = _build_service(profile=_build_profile(), returns=[ret])
        issues = service.check_readiness(1, 2025)
        codes = {i.code for i in issues}
        self.assertIn("UNFILED_RETURNS", codes)

    def test_readiness_clean_for_full_year_filed(self) -> None:
        obligations = [
            _build_obligation(date(2025, m, 1), date(2025, m, 28), id=m)
            for m in range(1, 13)
        ]
        returns = [
            _build_return(
                date(2025, m, 1),
                date(2025, m, 28),
                id=m,
                obligation_id=m,
                status_code=RETURN_STATUS_FILED,
            )
            for m in range(1, 13)
        ]
        service = _build_service(
            profile=_build_profile(),
            obligations=obligations,
            returns=returns,
        )
        issues = service.check_readiness(1, 2025)
        # No errors; ideally no warnings for this happy path.
        self.assertFalse(any(i.severity == "error" for i in issues))

    # ── Workbook generation ───────────────────────────────────────

    def test_generate_writes_expected_sheets(self) -> None:
        service = _build_service(profile=_build_profile())
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertTrue(os.path.isfile(self.output_path))
        self.assertEqual(
            result.sheets_written,
            ("Company Profile", "VAT Summary", "VAT Detail", "Payments", "Readiness"),
        )

    def test_generate_excludes_data_outside_year(self) -> None:
        in_year = _build_obligation(date(2025, 6, 1), date(2025, 6, 30), id=1)
        out_of_year = _build_obligation(date(2024, 12, 1), date(2024, 12, 31), id=2)
        service = _build_service(
            profile=_build_profile(),
            obligations=[in_year, out_of_year],
        )
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertEqual(result.obligation_count, 1)

    def test_generate_creates_missing_output_directory(self) -> None:
        nested = os.path.join(self._tmpdir.name, "exports", "2025", "dsf.xlsx")
        service = _build_service(profile=_build_profile())
        result = service.generate(1, GenerateDSFExportCommand(2025, nested))
        self.assertTrue(os.path.isfile(nested))
        self.assertEqual(result.output_path, nested)

    def test_generate_includes_returns_and_payments_in_year(self) -> None:
        ret = _build_return(
            date(2025, 1, 1),
            date(2025, 1, 31),
            id=10,
            obligation_id=100,
            status_code=RETURN_STATUS_FILED,
            total_due_amount=Decimal("150000.00"),
            total_paid_amount=Decimal("150000.00"),
            filed_at=datetime(2025, 2, 14, 10, 0, 0),
            lines=[
                _build_line(VAT_BOX_OUTPUT_TAX, Decimal("200000.00"), sort_order=1),
                _build_line(
                    VAT_BOX_INPUT_TAX_DEDUCTIBLE,
                    Decimal("50000.00"),
                    sort_order=2,
                ),
                _build_line(VAT_BOX_NET_VAT_DUE, Decimal("150000.00"), sort_order=3),
            ],
        )
        payment = _build_payment(
            date(2025, 2, 14), Decimal("150000.00"), id=1, tax_return_id=10
        )
        service = _build_service(
            profile=_build_profile(),
            returns=[ret],
            payments=[payment],
        )
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertEqual(result.return_count, 1)
        self.assertEqual(result.payment_count, 1)

        # Verify workbook content
        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        self.assertIn("VAT Detail", wb.sheetnames)
        ws = wb["VAT Detail"]
        # Header row + 3 lines
        self.assertEqual(ws.max_row, 4)
        # Box codes from sorted order
        self.assertEqual(ws.cell(row=2, column=4).value, VAT_BOX_OUTPUT_TAX)

    def test_generate_writes_company_profile_fields(self) -> None:
        profile = _build_profile(niu="P987654321B")
        service = _build_service(profile=profile)
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Company Profile"]
        rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(1, ws.max_row + 1)}
        self.assertEqual(rows.get("NIU"), "P987654321B")
        self.assertEqual(rows.get("Fiscal year"), "2025")

    def test_generate_handles_missing_profile_without_crash(self) -> None:
        service = _build_service(profile=None)
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertTrue(result.has_blocking_issues)
        self.assertTrue(os.path.isfile(self.output_path))

    def test_readiness_sheet_lists_issues(self) -> None:
        service = _build_service(profile=None)
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Readiness"]
        codes = {ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)}
        self.assertIn("NO_TAX_PROFILE", codes)


class DSFExportRegimeConditionalFichesTests(unittest.TestCase):
    """Phase 4 — regime-conditional DSF fiche layouts."""

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.output_path = os.path.join(self._tmpdir.name, "dsf.xlsx")

    def _generate_with_form(
        self,
        *,
        dsf_form_code: str,
        tax_regime_code: str,
        obligations: list | None = None,
        returns: list | None = None,
        payments: list | None = None,
    ):
        profile = _build_profile(
            dsf_form_code=dsf_form_code,
            tax_regime_code=tax_regime_code,
        )
        service = _build_service(
            profile=profile,
            obligations=obligations or [],
            returns=returns or [],
            payments=payments or [],
        )
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        return result, wb

    # ── DSF_REAL — full fiche set ────────────────────────────────

    def test_real_form_emits_full_fiche_set(self) -> None:
        result, wb = self._generate_with_form(
            dsf_form_code="DSF_REAL", tax_regime_code="REAL"
        )
        self.assertEqual(result.dsf_form_applied, "DSF_REAL")
        self.assertEqual(result.tax_regime_applied, "REAL")
        for sheet in (
            "Fiche R1 - Identity",
            "Fiche R2 - Revenue",
            "Fiche R3 - Balance Sheet",
            "Fiche R4 - Income Statement",
            "Fiche - CIT Summary",
        ):
            self.assertIn(sheet, wb.sheetnames, f"missing sheet: {sheet}")
        # Base sheets still present
        for sheet in (
            "Company Profile",
            "VAT Summary",
            "VAT Detail",
            "Payments",
            "Readiness",
        ):
            self.assertIn(sheet, wb.sheetnames)

    def test_real_fiche_r2_aggregates_vat_totals(self) -> None:
        ret = _build_return(
            date(2025, 1, 1),
            date(2025, 1, 31),
            id=1,
            obligation_id=1,
            status_code=RETURN_STATUS_FILED,
            total_due_amount=Decimal("150000.00"),
            total_paid_amount=Decimal("150000.00"),
            lines=[
                _build_line("BASE_19_25", Decimal("1000000.00"), sort_order=1),
                _build_line(VAT_BOX_OUTPUT_TAX, Decimal("192500.00"), sort_order=2),
            ],
        )
        result, wb = self._generate_with_form(
            dsf_form_code="DSF_REAL",
            tax_regime_code="REAL",
            returns=[ret],
        )
        ws = wb["Fiche R2 - Revenue"]
        # Header + at least 6 data rows (3 totals + 3 placeholders)
        self.assertGreaterEqual(ws.max_row, 7)
        codes = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        self.assertIn("R2-CA-TOTAL", codes)
        self.assertIn("R2-VAT-DUE", codes)
        self.assertIn("R2-VAT-PAID", codes)

        # Find the row whose code is R2-VAT-DUE and verify amount
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "R2-VAT-DUE":
                self.assertEqual(ws.cell(row=r, column=3).value, 150000.0)
                break
        else:
            self.fail("R2-VAT-DUE row missing")

    def test_real_fiche_r3_balance_sheet_carries_syscohada_structure(self) -> None:
        _, wb = self._generate_with_form(
            dsf_form_code="DSF_REAL", tax_regime_code="REAL"
        )
        ws = wb["Fiche R3 - Balance Sheet"]
        refs = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        # Sample of canonical SYSCOHADA refs
        for ref in ("AZ", "BZ", "CZ", "CP", "DZ", "EZ"):
            self.assertIn(ref, refs, f"missing SYSCOHADA ref: {ref}")

    def test_real_fiche_cit_summary_aggregates_cit_obligations(self) -> None:
        cit_obl = SimpleNamespace(
            id=200,
            company_id=1,
            tax_type_code="CIT_INSTALLMENT",
            period_start=date(2025, 1, 1),
            period_end=date(2025, 3, 31),
            due_date=date(2025, 4, 15),
            status_code="FILED",
        )
        cit_ret = SimpleNamespace(
            id=20,
            company_id=1,
            obligation_id=200,
            tax_type_code="CIT_INSTALLMENT",
            period_start=date(2025, 1, 1),
            period_end=date(2025, 3, 31),
            status_code=RETURN_STATUS_FILED,
            total_due_amount=Decimal("500000.00"),
            total_paid_amount=Decimal("500000.00"),
            filed_at=datetime(2025, 4, 14, 10, 0, 0),
            otp_reference=None,
            external_reference=None,
            lines=[],
            payments=[],
        )
        _, wb = self._generate_with_form(
            dsf_form_code="DSF_REAL",
            tax_regime_code="REAL",
            obligations=[cit_obl],
            returns=[cit_ret],
        )
        ws = wb["Fiche - CIT Summary"]
        labels_col_e = {ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)}
        self.assertIn("TOTAL", labels_col_e)
        self.assertIn("NET BALANCE", labels_col_e)

    # ── DSF_SIMPLIFIED — reduced fiche set ───────────────────────

    def test_simplified_form_emits_simplified_fiches(self) -> None:
        result, wb = self._generate_with_form(
            dsf_form_code="DSF_SIMPLIFIED", tax_regime_code="SIMPLIFIED"
        )
        self.assertEqual(result.dsf_form_applied, "DSF_SIMPLIFIED")
        self.assertIn("Fiche R1 - Identity", wb.sheetnames)
        self.assertIn("Fiche R2 - Revenue", wb.sheetnames)
        self.assertIn("Fiche - Simplified P&L", wb.sheetnames)
        self.assertIn("Fiche - CIT Summary", wb.sheetnames)
        # Full-regime-only sheets are NOT present
        self.assertNotIn("Fiche R3 - Balance Sheet", wb.sheetnames)
        self.assertNotIn("Fiche R4 - Income Statement", wb.sheetnames)

    def test_simplified_pl_uses_compact_row_set(self) -> None:
        _, wb = self._generate_with_form(
            dsf_form_code="DSF_SIMPLIFIED", tax_regime_code="SIMPLIFIED"
        )
        ws = wb["Fiche - Simplified P&L"]
        refs = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        for ref in ("S1", "S5", "S7"):
            self.assertIn(ref, refs)

    # ── DSF_LIBERATORY — minimal IGS summary ─────────────────────

    def test_liberatory_form_emits_summary_only(self) -> None:
        payment = _build_payment(
            date(2025, 6, 30), Decimal("75000.00"), id=1, tax_return_id=None
        )
        result, wb = self._generate_with_form(
            dsf_form_code="DSF_LIBERATORY",
            tax_regime_code="LIBERATORY",
            payments=[payment],
        )
        self.assertEqual(result.dsf_form_applied, "DSF_LIBERATORY")
        self.assertIn("Fiche R1 - Identity", wb.sheetnames)
        self.assertIn("Fiche - Liberatory Summary", wb.sheetnames)
        # Real/Simplified sheets are NOT present
        self.assertNotIn("Fiche R2 - Revenue", wb.sheetnames)
        self.assertNotIn("Fiche R3 - Balance Sheet", wb.sheetnames)
        self.assertNotIn("Fiche R4 - Income Statement", wb.sheetnames)
        self.assertNotIn("Fiche - Simplified P&L", wb.sheetnames)
        self.assertNotIn("Fiche - CIT Summary", wb.sheetnames)

        ws = wb["Fiche - Liberatory Summary"]
        # Verify the payments roll-up landed
        rows_by_label = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertEqual(
            rows_by_label.get("Total taxes paid / Total impôts payés"), 75000.0
        )
        self.assertEqual(
            rows_by_label.get("Number of payments / Nombre de paiements"), 1
        )

    # ── NONE / unknown / unset ───────────────────────────────────

    def test_none_form_emits_only_base_sheets(self) -> None:
        result, wb = self._generate_with_form(
            dsf_form_code="NONE", tax_regime_code="LIBERATORY"
        )
        self.assertIsNone(result.dsf_form_applied)
        self.assertEqual(
            result.sheets_written,
            ("Company Profile", "VAT Summary", "VAT Detail", "Payments", "Readiness"),
        )
        self.assertNotIn("Fiche R1 - Identity", wb.sheetnames)

    def test_unknown_form_code_falls_back_to_base_sheets(self) -> None:
        # Legacy / unknown codes should not crash — fiches simply are
        # not produced. The base export still succeeds.
        result, wb = self._generate_with_form(
            dsf_form_code="DSF_LEGACY_UNKNOWN", tax_regime_code="REAL"
        )
        self.assertIsNone(result.dsf_form_applied)
        self.assertNotIn("Fiche R1 - Identity", wb.sheetnames)


class DSFExportFicheBalancePopulationTests(unittest.TestCase):
    """Phase 4 deferred — Fiche R3/R4 amount population from posted GL."""

    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.output_path = os.path.join(self._tmpdir.name, "dsf.xlsx")

    def _make_balance_sheet_service(self, ref_to_amounts: dict) -> MagicMock:
        asset_lines = []
        liability_lines = []
        for ref, amounts in ref_to_amounts.items():
            line = SimpleNamespace(
                reference_code=ref,
                gross_amount=amounts.get("gross"),
                contra_amount=amounts.get("contra"),
                net_amount=amounts.get("net"),
            )
            # Place CP/DZ/EZ on liabilities, others on assets
            if ref in ("CA", "CB", "CD", "CK", "CP", "DA", "DP", "DH", "DI", "DZ", "EZ"):
                liability_lines.append(line)
            else:
                asset_lines.append(line)
        dto = SimpleNamespace(
            asset_lines=tuple(asset_lines),
            liability_lines=tuple(liability_lines),
        )
        svc = MagicMock()
        svc.get_statement.return_value = dto
        return svc

    def _make_income_statement_service(self, code_to_amount: dict) -> MagicMock:
        lines = tuple(
            SimpleNamespace(code=code, signed_amount=amount)
            for code, amount in code_to_amount.items()
        )
        dto = SimpleNamespace(lines=lines)
        svc = MagicMock()
        svc.get_statement.return_value = dto
        return svc

    def test_fiche_r3_populated_from_balance_sheet_service(self) -> None:
        bs = self._make_balance_sheet_service(
            {
                "AI": {"gross": Decimal("5000000.00"), "contra": Decimal("1200000.00"), "net": Decimal("3800000.00")},
                "BG": {"gross": Decimal("750000.00"), "contra": Decimal("0.00"), "net": Decimal("750000.00")},
                "CP": {"gross": Decimal("4550000.00"), "contra": Decimal("0.00"), "net": Decimal("4550000.00")},
            }
        )
        profile = _build_profile(dsf_form_code="DSF_REAL", tax_regime_code="REAL")
        service = _build_service(
            profile=profile, ohada_balance_sheet_service=bs
        )
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Fiche R3 - Balance Sheet"]
        # Find the AI row and assert the gross/contra/net populated
        ai_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "AI":
                ai_row = r
                break
        self.assertIsNotNone(ai_row, "AI row missing on Fiche R3")
        self.assertEqual(ws.cell(row=ai_row, column=3).value, 5000000.0)
        self.assertEqual(ws.cell(row=ai_row, column=4).value, 1200000.0)
        self.assertEqual(ws.cell(row=ai_row, column=5).value, 3800000.0)
        # Prior-period column always blank
        self.assertIsNone(ws.cell(row=ai_row, column=6).value)

    def test_fiche_r4_income_statement_populated(self) -> None:
        is_svc = self._make_income_statement_service(
            {
                "TA": Decimal("8000000.00"),
                "RA": Decimal("3200000.00"),
                "RI": Decimal("950000.00"),
            }
        )
        profile = _build_profile(dsf_form_code="DSF_REAL", tax_regime_code="REAL")
        service = _build_service(
            profile=profile, ohada_income_statement_service=is_svc
        )
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Fiche R4 - Income Statement"]
        # Map ref -> Year-N value
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertEqual(values.get("TA"), 8000000.0)
        self.assertEqual(values.get("RA"), 3200000.0)
        self.assertEqual(values.get("RI"), 950000.0)
        # Refs not provided remain blank
        self.assertIsNone(values.get("TC"))

    def test_fiche_r4_simplified_uses_amounts(self) -> None:
        is_svc = self._make_income_statement_service(
            {"S1": Decimal("1500000.00"), "S7": Decimal("250000.00")}
        )
        profile = _build_profile(
            dsf_form_code="DSF_SIMPLIFIED", tax_regime_code="SIMPLIFIED"
        )
        service = _build_service(
            profile=profile, ohada_income_statement_service=is_svc
        )
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws = wb["Fiche - Simplified P&L"]
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=3).value
            for r in range(2, ws.max_row + 1)
        }
        self.assertEqual(values.get("S1"), 1500000.0)
        self.assertEqual(values.get("S7"), 250000.0)

    def test_balance_sheet_service_failure_records_warning(self) -> None:
        from seeker_accounting.platform.exceptions import ValidationError

        bs = MagicMock()
        bs.get_statement.side_effect = ValidationError("no chart coverage")
        profile = _build_profile(dsf_form_code="DSF_REAL", tax_regime_code="REAL")
        service = _build_service(
            profile=profile, ohada_balance_sheet_service=bs
        )
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        codes = {issue.code for issue in result.readiness_issues}
        self.assertIn("FICHE_R3_AMOUNTS_UNAVAILABLE", codes)

    def test_no_services_injected_leaves_amounts_blank(self) -> None:
        profile = _build_profile(dsf_form_code="DSF_REAL", tax_regime_code="REAL")
        service = _build_service(profile=profile)
        service.generate(1, GenerateDSFExportCommand(2025, self.output_path))

        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        ws_r3 = wb["Fiche R3 - Balance Sheet"]
        ws_r4 = wb["Fiche R4 - Income Statement"]
        # Sample a few cells — all blank
        for r in range(2, ws_r3.max_row + 1):
            for c in (3, 4, 5, 6):
                self.assertIsNone(ws_r3.cell(row=r, column=c).value)
        for r in range(2, ws_r4.max_row + 1):
            self.assertIsNone(ws_r4.cell(row=r, column=3).value)


if __name__ == "__main__":
    unittest.main()
