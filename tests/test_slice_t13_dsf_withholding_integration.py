"""Slice T13 — DSF export integration with the WHT certificate register.

Verifies that when a withholding-tax certificate repository factory
is wired into ``DSFExportService``, the produced workbook gains a
``Withholding Certificates`` sheet with the year's rows, and the
result DTO reports the correct ``withholding_certificate_count``.
"""

from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock

from seeker_accounting.modules.taxation.constants import (
    DSF_FORM_NONE,
    TAX_REGIME_REAL,
    WHT_DIRECTION_INBOUND,
    WHT_DIRECTION_OUTBOUND,
    WHT_STATUS_ISSUED,
    WHT_STATUS_RECEIVED,
)
from seeker_accounting.modules.taxation.dto.dsf_export_dto import (
    GenerateDSFExportCommand,
)
from seeker_accounting.modules.taxation.services.dsf_export_service import (
    DSFExportService,
)


class _FakePermissionService:
    def __init__(self, granted: set[str]) -> None:
        self._granted = granted

    def require_permission(self, permission_code: str) -> None:
        if permission_code not in self._granted:
            from seeker_accounting.platform.exceptions import PermissionDeniedError

            raise PermissionDeniedError(permission_code)

    def has_permission(self, permission_code: str) -> bool:
        return permission_code in self._granted


class _FakeUnitOfWork:
    def __init__(self) -> None:
        self.session = MagicMock()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False


def _build_profile():
    return SimpleNamespace(
        company_id=1,
        niu="P012345678901A",
        tax_center_code="DGE",
        taxpayer_segment_code="LARGE",
        tax_regime_code=TAX_REGIME_REAL,
        is_vat_liable=True,
        vat_effective_from=date(2024, 1, 1),
        cit_rate_profile_code="STANDARD",
        cit_installment_profile_code="QUARTERLY",
        sme_qualified_flag=False,
        dsf_form_code=DSF_FORM_NONE,
        dsf_submission_mode_code="ELECTRONIC",
        otp_enabled_flag=True,
        default_withholding_applicable_flag=False,
    )


def _build_certificate(
    *,
    cid: int,
    direction: str,
    status: str,
    number: str,
    cert_date: date,
    base: Decimal,
    tax: Decimal,
):
    return SimpleNamespace(
        id=cid,
        company_id=1,
        direction=direction,
        counterparty_kind="CUSTOMER" if direction == WHT_DIRECTION_INBOUND else "SUPPLIER",
        counterparty_id=None,
        counterparty_name=f"Party {cid}",
        counterparty_niu=None,
        tax_code_id=10,
        certificate_number=number,
        certificate_date=cert_date,
        taxable_base=base,
        tax_amount=tax,
        status_code=status,
        notes=None,
    )


class DSFExportWithholdingIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmpdir.cleanup)
        self.output_path = os.path.join(self._tmpdir.name, "dsf_2025.xlsx")

    def _build_service(self, *, wht_rows):
        uow = _FakeUnitOfWork()

        profile_repo = MagicMock()
        profile_repo.get_by_company.return_value = _build_profile()

        obligation_repo = MagicMock()
        obligation_repo.list_by_company.return_value = []

        return_repo = MagicMock()
        return_repo.list_by_company.return_value = []
        return_repo.get_by_id.return_value = None

        payment_repo = MagicMock()
        payment_repo.list_by_company.return_value = []

        company_repo = MagicMock()
        company_repo.get_by_id.return_value = SimpleNamespace(id=1, name="Acme")

        wht_repo = MagicMock()
        wht_repo.list_by_company.return_value = wht_rows

        return DSFExportService(
            unit_of_work_factory=lambda: uow,
            app_context=SimpleNamespace(current_user_id=42),
            company_tax_profile_repository_factory=lambda s: profile_repo,
            tax_obligation_repository_factory=lambda s: obligation_repo,
            tax_return_repository_factory=lambda s: return_repo,
            tax_payment_repository_factory=lambda s: payment_repo,
            company_repository_factory=lambda s: company_repo,
            permission_service=_FakePermissionService(
                {DSFExportService.PERMISSION_EXPORT}
            ),
            audit_service=None,
            withholding_tax_certificate_repository_factory=lambda s: wht_repo,
        ), wht_repo

    def test_optional_factory_default_omits_sheet_and_reports_zero(self) -> None:
        # Build a service WITHOUT the WHT factory — historical behavior
        uow = _FakeUnitOfWork()
        profile_repo = MagicMock()
        profile_repo.get_by_company.return_value = _build_profile()
        obligation_repo = MagicMock()
        obligation_repo.list_by_company.return_value = []
        return_repo = MagicMock()
        return_repo.list_by_company.return_value = []
        return_repo.get_by_id.return_value = None
        payment_repo = MagicMock()
        payment_repo.list_by_company.return_value = []
        company_repo = MagicMock()
        company_repo.get_by_id.return_value = SimpleNamespace(id=1, name="Acme")

        service = DSFExportService(
            unit_of_work_factory=lambda: uow,
            app_context=SimpleNamespace(current_user_id=42),
            company_tax_profile_repository_factory=lambda s: profile_repo,
            tax_obligation_repository_factory=lambda s: obligation_repo,
            tax_return_repository_factory=lambda s: return_repo,
            tax_payment_repository_factory=lambda s: payment_repo,
            company_repository_factory=lambda s: company_repo,
            permission_service=_FakePermissionService(
                {DSFExportService.PERMISSION_EXPORT}
            ),
            audit_service=None,
        )
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertNotIn("Withholding Certificates", result.sheets_written)
        self.assertEqual(result.withholding_certificate_count, 0)

    def test_wired_factory_writes_sheet_for_year_rows(self) -> None:
        rows = [
            _build_certificate(
                cid=1,
                direction=WHT_DIRECTION_INBOUND,
                status=WHT_STATUS_RECEIVED,
                number="IN-001",
                cert_date=date(2025, 3, 15),
                base=Decimal("1000.00"),
                tax=Decimal("22.00"),
            ),
            _build_certificate(
                cid=2,
                direction=WHT_DIRECTION_OUTBOUND,
                status=WHT_STATUS_ISSUED,
                number="OUT-001",
                cert_date=date(2025, 6, 1),
                base=Decimal("2000.00"),
                tax=Decimal("110.00"),
            ),
        ]
        service, wht_repo = self._build_service(wht_rows=rows)
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertIn("Withholding Certificates", result.sheets_written)
        self.assertEqual(result.withholding_certificate_count, 2)

        # Repo was called with the fiscal-year window
        wht_repo.list_by_company.assert_called_once()
        kwargs = wht_repo.list_by_company.call_args.kwargs
        self.assertEqual(kwargs["date_from"], date(2025, 1, 1))
        self.assertEqual(kwargs["date_to"], date(2025, 12, 31))

        # Sheet content present
        import openpyxl

        wb = openpyxl.load_workbook(self.output_path)
        self.assertIn("Withholding Certificates", wb.sheetnames)
        ws = wb["Withholding Certificates"]
        self.assertEqual(ws.cell(row=1, column=1).value, "Direction")
        self.assertEqual(ws.cell(row=1, column=11).value, "Notes")
        # 2 data rows under the header
        directions = {ws.cell(row=r, column=1).value for r in (2, 3)}
        self.assertEqual(directions, {WHT_DIRECTION_INBOUND, WHT_DIRECTION_OUTBOUND})

    def test_empty_rows_skip_sheet_but_count_is_zero(self) -> None:
        service, _ = self._build_service(wht_rows=[])
        result = service.generate(
            1, GenerateDSFExportCommand(2025, self.output_path)
        )
        self.assertNotIn("Withholding Certificates", result.sheets_written)
        self.assertEqual(result.withholding_certificate_count, 0)


if __name__ == "__main__":
    unittest.main()
