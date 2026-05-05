"""Slice T22-T25 tests — dashboard / audit trail / PDF export / DSF P&L+BS.

Service-layer tests for the four taxation services added in slices
T22 (TaxDashboardService), T23 (TaxAuditTrailService), T24
(TaxReturnPDFExportService), and T25 (DSF expanded P&L/BS sheets).

These tests follow the mock-based pattern established in earlier
taxation slices: unit of work, repositories, and permission service
are stubbed so we exercise the service surface in isolation.
"""

from __future__ import annotations

import os
import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import MagicMock

from seeker_accounting.db import model_registry  # noqa: F401
from seeker_accounting.modules.taxation.constants import (
    OBLIGATION_STATUS_CANCELLED,
    OBLIGATION_STATUS_OPEN,
    OBLIGATION_STATUS_OVERDUE,
    OBLIGATION_STATUS_PAID,
    RETURN_STATUS_DRAFT,
    RETURN_STATUS_FILED,
    TAX_TYPE_CIT_INSTALLMENT,
    TAX_TYPE_VAT,
    WHT_DIRECTION_INBOUND,
    WHT_DIRECTION_OUTBOUND,
)
from seeker_accounting.modules.taxation.dto.tax_compliance_dto import (
    ExportTaxReturnPDFCommand,
    TaxAuditFilterDTO,
)
from seeker_accounting.modules.taxation.services.tax_audit_trail_service import (
    TaxAuditTrailService,
)
from seeker_accounting.modules.taxation.services.tax_dashboard_service import (
    TaxDashboardService,
)
from seeker_accounting.modules.taxation.services.tax_return_pdf_export_service import (
    TaxReturnPDFExportService,
)
from seeker_accounting.platform.exceptions import (
    NotFoundError,
    PermissionDeniedError,
    ValidationError,
)


# ─── Test helpers ────────────────────────────────────────────────────


class _FakeUnitOfWork:
    def __init__(self) -> None:
        self.session = MagicMock(name="Session")
        self.committed = False

    def __enter__(self) -> "_FakeUnitOfWork":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        return None

    def commit(self) -> None:
        self.committed = True


class _FakePermissionService:
    def __init__(self, granted: set[str]) -> None:
        self._granted = granted

    def require_permission(self, code: str) -> None:
        if code not in self._granted:
            raise PermissionDeniedError(f"Missing permission: {code}")


def _ob(
    *,
    id_: int,
    tax_type_code: str,
    period_start: date,
    period_end: date,
    due_date: date,
    status_code: str = OBLIGATION_STATUS_OPEN,
):
    return SimpleNamespace(
        id=id_,
        company_id=1,
        tax_type_code=tax_type_code,
        period_start=period_start,
        period_end=period_end,
        due_date=due_date,
        status_code=status_code,
        notes=None,
        created_at=None,
        updated_at=None,
    )


def _ret(
    *,
    id_: int,
    tax_type_code: str,
    period_end: date,
    status_code: str,
    journal_entry_id: int | None = None,
    total_due: Decimal = Decimal("0.00"),
):
    return SimpleNamespace(
        id=id_,
        company_id=1,
        obligation_id=10 + id_,
        tax_type_code=tax_type_code,
        period_start=date(period_end.year, 1, 1),
        period_end=period_end,
        status_code=status_code,
        total_due_amount=total_due,
        total_paid_amount=Decimal("0.00"),
        filed_at=None,
        otp_reference=None,
        external_reference=None,
        notes=None,
        prepared_by_user_id=None,
        journal_entry_id=journal_entry_id,
        lines=[],
    )


def _pay(*, id_: int, payment_date: date, amount: Decimal):
    return SimpleNamespace(
        id=id_,
        company_id=1,
        tax_return_id=1,
        payment_date=payment_date,
        amount=amount,
        payment_method_code="BANK",
        reference=None,
        notes=None,
        journal_entry_id=None,
        recorded_by_user_id=None,
    )


def _wht(*, id_: int, direction: str, amount: Decimal):
    return SimpleNamespace(
        id=id_,
        company_id=1,
        direction=direction,
        certificate_number=f"WHT-{id_}",
        certificate_date=date(2026, 6, 1),
        counterparty_kind="SUPPLIER",
        counterparty_name="Acme",
        counterparty_niu=None,
        tax_code_id=1,
        status_code="ACTIVE",
        taxable_base=amount * 10,
        tax_amount=amount,
        notes=None,
    )


# ─── T22 dashboard tests ──────────────────────────────────────────────


def _build_dashboard_service(
    *,
    granted: set[str] | None = None,
    obligations: list | None = None,
    returns: list | None = None,
    payments: list | None = None,
    wht_inbound: list | None = None,
    wht_outbound: list | None = None,
    company_exists: bool = True,
):
    if granted is None:
        granted = {TaxDashboardService.PERMISSION_VIEW}

    uow = _FakeUnitOfWork()

    ob_repo = MagicMock()
    ob_repo.list_by_company.return_value = obligations or []

    ret_repo = MagicMock()
    ret_repo.list_by_company.return_value = returns or []

    pay_repo = MagicMock()
    pay_repo.list_by_company.return_value = payments or []

    wht_repo = MagicMock()

    def _wht_list(company_id, *, direction, date_from, date_to):
        return (wht_inbound or []) if direction == WHT_DIRECTION_INBOUND else (
            wht_outbound or []
        )

    wht_repo.list_by_company.side_effect = _wht_list

    company_repo = MagicMock()
    company_repo.get_by_id.return_value = (
        SimpleNamespace(id=1) if company_exists else None
    )

    service = TaxDashboardService(
        unit_of_work_factory=lambda: uow,
        app_context=SimpleNamespace(current_user_id=42),
        tax_obligation_repository_factory=lambda s: ob_repo,
        tax_return_repository_factory=lambda s: ret_repo,
        tax_payment_repository_factory=lambda s: pay_repo,
        withholding_tax_certificate_repository_factory=lambda s: wht_repo,
        company_repository_factory=lambda s: company_repo,
        permission_service=_FakePermissionService(granted),
    )
    return service


class DashboardSnapshotTests(unittest.TestCase):
    def test_basic_aggregate_counts_and_totals(self) -> None:
        as_of = date(2026, 6, 15)
        obligations = [
            _ob(
                id_=1,
                tax_type_code=TAX_TYPE_VAT,
                period_start=date(2026, 1, 1),
                period_end=date(2026, 1, 31),
                due_date=date(2026, 2, 15),
                status_code=OBLIGATION_STATUS_PAID,
            ),
            _ob(
                id_=2,
                tax_type_code=TAX_TYPE_VAT,
                period_start=date(2026, 5, 1),
                period_end=date(2026, 5, 31),
                due_date=date(2026, 5, 15),  # before as_of -> overdue
                status_code=OBLIGATION_STATUS_OPEN,
            ),
            _ob(
                id_=3,
                tax_type_code=TAX_TYPE_VAT,
                period_start=date(2026, 6, 1),
                period_end=date(2026, 6, 30),
                due_date=date(2026, 7, 15),
                status_code=OBLIGATION_STATUS_OPEN,
            ),
            _ob(
                id_=4,
                tax_type_code=TAX_TYPE_CIT_INSTALLMENT,
                period_start=date(2026, 1, 1),
                period_end=date(2026, 3, 31),
                due_date=date(2026, 4, 15),
                status_code=OBLIGATION_STATUS_CANCELLED,
            ),
        ]
        returns = [
            _ret(
                id_=1,
                tax_type_code=TAX_TYPE_VAT,
                period_end=date(2026, 1, 31),
                status_code=RETURN_STATUS_FILED,
                journal_entry_id=99,
                total_due=Decimal("100000"),
            ),
            _ret(
                id_=2,
                tax_type_code=TAX_TYPE_VAT,
                period_end=date(2026, 5, 31),
                status_code=RETURN_STATUS_FILED,
                journal_entry_id=None,  # unsettled VAT
                total_due=Decimal("50000"),
            ),
            _ret(
                id_=3,
                tax_type_code=TAX_TYPE_VAT,
                period_end=date(2026, 6, 30),
                status_code=RETURN_STATUS_DRAFT,
            ),
        ]
        payments = [
            _pay(id_=1, payment_date=date(2026, 2, 20), amount=Decimal("90000")),
            _pay(id_=2, payment_date=date(2026, 6, 1), amount=Decimal("10000")),
        ]
        wht_in = [_wht(id_=1, direction=WHT_DIRECTION_INBOUND, amount=Decimal("500"))]
        wht_out = [
            _wht(id_=2, direction=WHT_DIRECTION_OUTBOUND, amount=Decimal("250")),
            _wht(id_=3, direction=WHT_DIRECTION_OUTBOUND, amount=Decimal("100")),
        ]

        service = _build_dashboard_service(
            obligations=obligations,
            returns=returns,
            payments=payments,
            wht_inbound=wht_in,
            wht_outbound=wht_out,
        )
        snap = service.get_dashboard(1, 2026, as_of_date=as_of)

        self.assertEqual(snap.fiscal_year, 2026)
        self.assertEqual(snap.total_obligations, 4)
        self.assertEqual(snap.paid_obligations, 1)
        self.assertEqual(snap.overdue_obligations, 1)  # ob 2 (due May 15 < June 15)
        self.assertEqual(snap.open_obligations, 1)  # ob 3
        self.assertEqual(snap.cancelled_obligations, 1)  # ob 4
        self.assertEqual(snap.returns_draft, 1)
        self.assertEqual(snap.returns_filed, 2)
        self.assertEqual(snap.returns_settled, 1)
        self.assertEqual(snap.returns_filed_unsettled_vat, 1)
        self.assertEqual(snap.total_payments_ytd, Decimal("100000"))
        self.assertEqual(snap.total_due_filed_returns_ytd, Decimal("150000"))
        self.assertEqual(snap.wht_inbound_total_ytd, Decimal("500"))
        self.assertEqual(snap.wht_outbound_total_ytd, Decimal("350"))
        # Upcoming should include open + overdue (sorted by due_date asc)
        self.assertEqual(len(snap.upcoming), 2)
        self.assertEqual(snap.upcoming[0].obligation_id, 2)  # earlier due_date
        self.assertEqual(snap.upcoming[0].days_until_due, -31)
        self.assertEqual(snap.upcoming[1].obligation_id, 3)

    def test_rejects_year_out_of_range(self) -> None:
        service = _build_dashboard_service()
        with self.assertRaises(ValidationError):
            service.get_dashboard(1, 1999)

    def test_requires_permission(self) -> None:
        service = _build_dashboard_service(granted=set())
        with self.assertRaises(PermissionDeniedError):
            service.get_dashboard(1, 2026)

    def test_missing_company_raises_not_found(self) -> None:
        service = _build_dashboard_service(company_exists=False)
        with self.assertRaises(NotFoundError):
            service.get_dashboard(1, 2026)

    def test_per_tax_type_summary_groups_correctly(self) -> None:
        as_of = date(2026, 6, 15)
        obligations = [
            _ob(
                id_=1,
                tax_type_code=TAX_TYPE_VAT,
                period_start=date(2026, 1, 1),
                period_end=date(2026, 1, 31),
                due_date=date(2026, 2, 15),
                status_code=OBLIGATION_STATUS_PAID,
            ),
            _ob(
                id_=2,
                tax_type_code=TAX_TYPE_CIT_INSTALLMENT,
                period_start=date(2026, 1, 1),
                period_end=date(2026, 3, 31),
                due_date=date(2026, 4, 15),
                status_code=OBLIGATION_STATUS_OPEN,  # past as_of -> overdue
            ),
        ]
        service = _build_dashboard_service(obligations=obligations)
        snap = service.get_dashboard(1, 2026, as_of_date=as_of)
        per_type = {s.tax_type_code: s for s in snap.by_tax_type}
        self.assertEqual(per_type[TAX_TYPE_VAT].paid_count, 1)
        self.assertEqual(per_type[TAX_TYPE_CIT_INSTALLMENT].overdue_count, 1)


# ─── T23 audit trail tests ────────────────────────────────────────────


class AuditTrailTests(unittest.TestCase):
    def _build(self, *, granted: set[str] | None = None):
        if granted is None:
            granted = {TaxAuditTrailService.PERMISSION_VIEW}
        audit_service = MagicMock()
        audit_service.list_events.return_value = []
        service = TaxAuditTrailService(
            audit_service=audit_service,
            permission_service=_FakePermissionService(granted),
        )
        return service, audit_service

    def test_delegates_to_audit_service_with_taxation_module_filter(self) -> None:
        service, audit = self._build()
        filt = TaxAuditFilterDTO(company_id=1, limit=50, offset=0)
        service.list_events(filt)
        audit.list_events.assert_called_once()
        kwargs = audit.list_events.call_args.kwargs
        self.assertEqual(kwargs["module_code"], "taxation")
        self.assertEqual(kwargs["limit"], 50)

    def test_requires_permission(self) -> None:
        service, _ = self._build(granted=set())
        with self.assertRaises(PermissionDeniedError):
            service.list_events(TaxAuditFilterDTO(company_id=1))

    def test_rejects_invalid_limit(self) -> None:
        service, _ = self._build()
        with self.assertRaises(ValidationError):
            service.list_events(TaxAuditFilterDTO(company_id=1, limit=0))
        with self.assertRaises(ValidationError):
            service.list_events(TaxAuditFilterDTO(company_id=1, limit=10000))

    def test_rejects_negative_offset(self) -> None:
        service, _ = self._build()
        with self.assertRaises(ValidationError):
            service.list_events(TaxAuditFilterDTO(company_id=1, offset=-1))

    def test_rejects_inverted_date_range(self) -> None:
        service, _ = self._build()
        with self.assertRaises(ValidationError):
            service.list_events(
                TaxAuditFilterDTO(
                    company_id=1,
                    from_date=datetime(2026, 12, 31),
                    to_date=datetime(2026, 1, 1),
                )
            )


# ─── T24 PDF export tests ─────────────────────────────────────────────


def _build_pdf_service(
    *,
    granted: set[str] | None = None,
    company_exists: bool = True,
    return_obj=None,
    print_engine=None,
):
    if granted is None:
        granted = {TaxReturnPDFExportService.PERMISSION_EXPORT}
    uow = _FakeUnitOfWork()
    company_repo = MagicMock()
    company_repo.get_by_id.return_value = (
        SimpleNamespace(
            id=1,
            display_name="Acme Test Co",
            legal_name="Acme Test Company SARL",
            tax_identifier="P123456789012X",
        )
        if company_exists
        else None
    )
    return_repo = MagicMock()
    return_repo.get_by_id.return_value = return_obj
    engine = print_engine or MagicMock()
    service = TaxReturnPDFExportService(
        unit_of_work_factory=lambda: uow,
        app_context=SimpleNamespace(
            current_user_id=42, current_user_display_name="Tester"
        ),
        tax_return_repository_factory=lambda s: return_repo,
        company_repository_factory=lambda s: company_repo,
        permission_service=_FakePermissionService(granted),
        print_engine=engine,
        audit_service=None,
    )
    return service, engine


class PDFExportTests(unittest.TestCase):
    def _make_return(self):
        return SimpleNamespace(
            id=42,
            company_id=1,
            tax_type_code=TAX_TYPE_VAT,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            status_code=RETURN_STATUS_FILED,
            total_due_amount=Decimal("123456.78"),
            total_paid_amount=Decimal("100000.00"),
            filed_at=datetime(2026, 6, 10, 10, 30),
            otp_reference="OTP-XYZ",
            external_reference="DGI-2026-05-001",
            notes="Filed via DGI portal.",
            lines=[
                SimpleNamespace(
                    id=1,
                    box_code="TAXABLE_SALES",
                    label="Taxable sales (HT)",
                    amount=Decimal("700000.00"),
                    sort_order=1,
                ),
                SimpleNamespace(
                    id=2,
                    box_code="VAT_OUTPUT",
                    label="Output VAT",
                    amount=Decimal("133000.00"),
                    sort_order=2,
                ),
            ],
        )

    def test_exports_pdf_calls_print_engine(self) -> None:
        engine = MagicMock()
        service, _ = _build_pdf_service(
            return_obj=self._make_return(), print_engine=engine
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "return.pdf")
            result = service.export(
                1, ExportTaxReturnPDFCommand(return_id=42, output_path=output)
            )
        self.assertEqual(result.return_id, 42)
        self.assertEqual(result.output_path, output)
        engine.render_pdf.assert_called_once()
        html_arg = engine.render_pdf.call_args.args[0]
        self.assertIn("Acme Test Co", html_arg)
        # DGI form structure must be present
        self.assertIn("Section 4", html_arg)
        self.assertIn("Turnover Realised", html_arg)
        self.assertIn("L17", html_arg)
        self.assertIn("L40", html_arg)
        # Aggregated values rendered into the form
        self.assertIn("700,000.00", html_arg)
        self.assertIn("133,000.00", html_arg)

    def test_rejects_non_pdf_extension(self) -> None:
        service, _ = _build_pdf_service(return_obj=self._make_return())
        with self.assertRaises(ValidationError):
            service.export(
                1, ExportTaxReturnPDFCommand(return_id=42, output_path="/tmp/x.txt")
            )

    def test_rejects_empty_output_path(self) -> None:
        service, _ = _build_pdf_service(return_obj=self._make_return())
        with self.assertRaises(ValidationError):
            service.export(
                1, ExportTaxReturnPDFCommand(return_id=42, output_path="   ")
            )

    def test_missing_return_raises_not_found(self) -> None:
        service, _ = _build_pdf_service(return_obj=None)
        with self.assertRaises(NotFoundError):
            service.export(
                1,
                ExportTaxReturnPDFCommand(
                    return_id=999, output_path="/tmp/x.pdf"
                ),
            )

    def test_missing_company_raises_not_found(self) -> None:
        service, _ = _build_pdf_service(
            company_exists=False, return_obj=self._make_return()
        )
        with self.assertRaises(NotFoundError):
            service.export(
                1, ExportTaxReturnPDFCommand(return_id=42, output_path="/tmp/x.pdf")
            )

    def test_requires_permission(self) -> None:
        service, _ = _build_pdf_service(
            granted=set(), return_obj=self._make_return()
        )
        with self.assertRaises(PermissionDeniedError):
            service.export(
                1, ExportTaxReturnPDFCommand(return_id=42, output_path="/tmp/x.pdf")
            )

    def test_handles_return_with_no_lines(self) -> None:
        engine = MagicMock()
        ret = self._make_return()
        ret.lines = []
        service, _ = _build_pdf_service(return_obj=ret, print_engine=engine)
        with tempfile.TemporaryDirectory() as tmpdir:
            service.export(
                1,
                ExportTaxReturnPDFCommand(
                    return_id=42,
                    output_path=os.path.join(tmpdir, "empty.pdf"),
                ),
            )
        html = engine.render_pdf.call_args.args[0]
        # Even with no aggregated lines, the DGI form structure is
        # still rendered (empty zeros under each section).
        self.assertIn("Section 4", html)
        self.assertIn("L17", html)
        self.assertIn("L40", html)


# ─── T25 DSF expanded sheets ──────────────────────────────────────────


class DSFExpandedSheetsTests(unittest.TestCase):
    """Validate that the DSF workbook now writes standalone P&L and
    Balance Sheet pages when the OHADA services are wired."""

    def test_writes_income_statement_and_balance_sheet_sheets(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")

        # Build a minimal in-memory DSF service exercising _write_workbook
        from seeker_accounting.modules.taxation.services.dsf_export_service import (
            DSFExportService,
        )

        # Construct fake full DTOs matching the OHADA service shape
        is_dto = SimpleNamespace(
            lines=(
                SimpleNamespace(
                    code="TA",
                    section_code="OP",
                    section_title="Operating",
                    label="Revenue",
                    signed_amount=Decimal("1000000"),
                    can_drilldown=True,
                ),
                SimpleNamespace(
                    code="TB",
                    section_code="OP",
                    section_title="Operating",
                    label="Cost of sales",
                    signed_amount=Decimal("-600000"),
                    can_drilldown=True,
                ),
            )
        )
        bs_dto = SimpleNamespace(
            asset_lines=(
                SimpleNamespace(
                    code="AB",
                    reference_code="AB",
                    label="Intangible assets",
                    gross_amount=Decimal("500000"),
                    contra_amount=Decimal("100000"),
                    net_amount=Decimal("400000"),
                ),
            ),
            liability_lines=(
                SimpleNamespace(
                    code="CA",
                    reference_code="CA",
                    label="Capital",
                    gross_amount=Decimal("400000"),
                    contra_amount=None,
                    net_amount=Decimal("400000"),
                ),
            ),
        )

        # Bypass real construction — call the bound writer through a
        # service instance with minimal init args.
        service = DSFExportService.__new__(DSFExportService)
        service._effective_dsf_form = lambda profile: None  # type: ignore[attr-defined]
        service._profile_rows = lambda profile, fy: [("Year", fy)]  # type: ignore[attr-defined]
        service._write_form_family_fiches = lambda **kw: ()  # type: ignore[attr-defined]

        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "dsf.xlsx")
            sheets = service._write_workbook(
                output_path=output,
                fiscal_year=2026,
                profile=SimpleNamespace(),
                obligations=[],
                returns=[],
                payments=[],
                issues=(),
                balance_sheet_amounts=None,
                income_statement_amounts=None,
                wht_certificates=None,
                balance_sheet_full_dto=bs_dto,
                income_statement_full_dto=is_dto,
            )
            self.assertIn("Income Statement", sheets)
            self.assertIn("Balance Sheet", sheets)
            # Reload and check content
            import openpyxl

            wb = openpyxl.load_workbook(output)
            self.assertIn("Income Statement", wb.sheetnames)
            self.assertIn("Balance Sheet", wb.sheetnames)
            is_ws = wb["Income Statement"]
            # header + 2 lines
            self.assertEqual(is_ws.max_row, 3)
            self.assertEqual(is_ws.cell(row=2, column=1).value, "TA")
            bs_ws = wb["Balance Sheet"]
            # header + asset(1) + divider + liab(1) = 4 rows
            self.assertEqual(bs_ws.max_row, 4)
            self.assertEqual(bs_ws.cell(row=2, column=1).value, "ASSET")
            self.assertEqual(bs_ws.cell(row=4, column=1).value, "LIAB/EQ")

    def test_omits_sheets_when_dtos_absent(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")

        from seeker_accounting.modules.taxation.services.dsf_export_service import (
            DSFExportService,
        )

        service = DSFExportService.__new__(DSFExportService)
        service._effective_dsf_form = lambda profile: None  # type: ignore[attr-defined]
        service._profile_rows = lambda profile, fy: [("Year", fy)]  # type: ignore[attr-defined]
        service._write_form_family_fiches = lambda **kw: ()  # type: ignore[attr-defined]

        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "dsf.xlsx")
            sheets = service._write_workbook(
                output_path=output,
                fiscal_year=2026,
                profile=SimpleNamespace(),
                obligations=[],
                returns=[],
                payments=[],
                issues=(),
                balance_sheet_amounts=None,
                income_statement_amounts=None,
                wht_certificates=None,
                balance_sheet_full_dto=None,
                income_statement_full_dto=None,
            )
            self.assertNotIn("Income Statement", sheets)
            self.assertNotIn("Balance Sheet", sheets)


if __name__ == "__main__":
    unittest.main()
